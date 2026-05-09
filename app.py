# import streamlit as st
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# import io
# from datetime import date

# st.set_page_config(page_title="PCB Assembly Cost Calculator", layout="wide")

# st.markdown("""
# <style>
#     .block-container { padding-top: 1.5rem; }
#     .section-header {
#         background: #1F4E79; color: white; padding: 8px 14px;
#         border-radius: 6px; font-weight: bold; margin: 18px 0 8px 0;
#     }
#     .sub-header {
#         background: #2E75B6; color: white; padding: 5px 12px;
#         border-radius: 4px; font-weight: 600; margin: 12px 0 6px 0; font-size: 0.9rem;
#     }
#     .result-box {
#         background: #EAF4FB; border: 1px solid #2E75B6;
#         border-radius: 6px; padding: 10px 16px; margin: 4px 0; color: #1a1a1a;
#     }
#     .total-box {
#         background: #2E75B6; border: 1px solid #1F4E79;
#         border-radius: 6px; padding: 10px 16px; margin: 6px 0;
#         font-weight: bold; color: #ffffff;
#     }
#     .grand-total {
#         background: #1F4E79; color: white; border-radius: 8px;
#         padding: 14px 20px; margin: 8px 0; font-size: 1.1rem; font-weight: bold;
#     }
#     .stNumberInput label, .stTextInput label { font-size: 0.85rem; }
# </style>
# """, unsafe_allow_html=True)

# st.markdown("## 🔧 PCB Assembly Cost Calculator")
# st.caption("Translated from: *Baugruppe LOHN* — exact formulas preserved from standard.xlsx")

# # ──────────────────────────────────────────────
# # SIDEBAR INPUTS
# # ──────────────────────────────────────────────
# with st.sidebar:
#     st.markdown("### ⚙️ General Settings")
#     qty_std  = st.number_input("Quantity Standard (G2)", value=50,  min_value=1, step=1)
#     qty_cust = st.number_input("Quantity Customer (H2)", value=100, min_value=1, step=1)
#     platine  = st.text_input("PCB / Platine name", value="PCB Assembly")
#     created_by = st.text_input("Created by", value="Ashok Kumar")
#     calc_date  = st.date_input("Date", value=date.today())

# # ──────────────────────────────────────────────
# # MAIN INPUTS — two column layout (Std | Cust)
# # ──────────────────────────────────────────────

# def num(label, val, key, fmt="%.2f", step=0.01, min_val=0.0):
#     return st.number_input(label, value=float(val), key=key, format=fmt, step=step, min_value=min_val)

# def num_int(label, val, key):
#     return st.number_input(label, value=int(val), key=key, step=1, min_value=0)

# # ── SECTION 1: SMD ──
# st.markdown('<div class="section-header">SECTION 1 — Labour Assembly (LOHN)</div>', unsafe_allow_html=True)
# st.markdown('<div class="sub-header">SMD Components</div>', unsafe_allow_html=True)

# col1, col2, col3 = st.columns([3, 1, 1])
# with col1:
#     st.markdown("SMD Components — Time per assembly *(Bauteile SMD)*")
# with col2:
#     C7 = num("Standard [Ct]", 5.5, "C7")
# with col3:
#     D7 = num("Customer [Ct]", 4.0, "D7")

# col1, col2 = st.columns([4, 1])
# with col1:
#     st.markdown("Number of SMD components")
# with col2:
#     E7 = num_int("Count", 120, "E7")

# st.markdown('<div class="sub-header">Soldering</div>', unsafe_allow_html=True)
# col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
# with col1:
#     st.markdown("Soldering 1-sided / 2-sided *(Lötung)*")
# with col2:
#     C8 = num("Standard [Ct]", 150, "C8", fmt="%.1f", step=1.0)
# with col3:
#     D8 = num("Customer [Ct]", 150, "D8", fmt="%.1f", step=1.0)
# with col4:
#     E8 = st.number_input("Sides (1 or 2)", value=2, min_value=1, max_value=2, step=1, key="E8")

# st.markdown('<div class="sub-header">FIX Costs per Side</div>', unsafe_allow_html=True)
# col1, col2, col3 = st.columns([3, 1, 1])
# with col1:
#     st.markdown("FIX Printer per side *(Drucker)*")
# with col2:
#     C9 = num("Standard €", 30, "C9")
# with col3:
#     D9 = num("Customer €", 30, "D9")

# col1, col2, col3 = st.columns([3, 1, 1])
# with col1:
#     st.markdown("FIX Machine per side *(Maschine)*")
# with col2:
#     C10 = num("Standard €", 30, "C10")
# with col3:
#     D10 = num("Customer €", 30, "D10")

# st.markdown('<div class="sub-header">SMD Feeder / Setup Types</div>', unsafe_allow_html=True)
# col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
# with col1:
#     st.markdown("SMD feeder types *(Verschiedene SMD-Bauteile)*")
# with col2:
#     C11 = num("Standard €/type", 6.5, "C11")
# with col3:
#     D11 = num("Customer €/type", 6.5, "D11")
# with col4:
#     E11 = num_int("No. of types", 33, "E11")

# st.markdown('<div class="sub-header">THT Components</div>', unsafe_allow_html=True)
# col1, col2, col3 = st.columns([3, 1, 1])
# with col1:
#     st.markdown("THT component types *(Verschiedene THT-Bauteile)*")
# with col2:
#     E12 = num_int("THT types", 0, "E12")

# col1, col2 = st.columns([4, 1])
# with col1:
#     st.markdown("Circuits per panel *(Einzelschaltungen pro Nutzen)*")
# with col2:
#     E13 = num_int("Count", 5, "E13")

# col1, col2, col3 = st.columns([3, 1, 1])
# with col1:
#     st.markdown("FIX Selective / Wave solder *(Selektiv oder Welle)*")
# with col2:
#     C14 = num("Standard €", 30, "C14")
# with col3:
#     D14 = num("Customer €", 30, "D14")

# col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
# with col1:
#     st.markdown("THT component cost *(Bauteile THT)*")
# with col2:
#     C15 = num("Standard [Ct]", 55, "C15", fmt="%.2f", step=0.5)
# with col3:
#     D15 = num("Customer [Ct]", 45, "D15", fmt="%.2f", step=0.5)
# with col4:
#     E15 = num_int("THT qty", 7, "E15")

# st.markdown('<div class="sub-header">Manual Soldering & QS</div>', unsafe_allow_html=True)
# col1, col2, col3 = st.columns([3, 1, 1])
# with col1:
#     st.markdown("Manual solder joints *(Einzellötstellen von Hand)*")
# with col2:
#     C16 = num("Standard count", 0, "C16", fmt="%.0f", step=1.0)
# with col3:
#     D16 = num("Customer count", 0, "D16", fmt="%.0f", step=1.0)

# col1, col2 = st.columns([4, 1])
# with col1:
#     st.markdown("Time per joint [min]")
# with col2:
#     E16 = num("Min/joint", 0.0, "E16")

# col1, col2 = st.columns([4, 1])
# with col1:
#     st.markdown("QS / Quality inspection time [min]")
# with col2:
#     F17 = num("QS min", 1.8, "F17")

# col1, col2, col3 = st.columns([3, 1, 1])
# with col1:
#     st.markdown("Hourly rate *(Faktor Stundensatz)*")
# with col2:
#     C18 = num("Standard €/h", 60, "C18", fmt="%.2f", step=1.0)
# with col3:
#     D18 = num("Customer €/h", 60, "D18", fmt="%.2f", step=1.0)

# # ── SECTION 2: Material ──
# st.markdown('<div class="section-header">SECTION 2 — Material</div>', unsafe_allow_html=True)
# col1, col2, col3 = st.columns([3, 1, 1])
# with col1:
#     st.markdown("Material cost per unit *(Materialkosten)*")
# with col2:
#     C25 = num("Standard €", 735, "C25", fmt="%.2f", step=1.0)
# with col3:
#     D25 = num("Customer €", 700, "D25", fmt="%.2f", step=1.0)

# col1, col2, col3 = st.columns([3, 1, 1])
# with col1:
#     st.markdown("Material markup % *(Aufschlag in %)*")
# with col2:
#     C26 = num("Standard %", 7, "C26", fmt="%.1f", step=0.5)
# with col3:
#     D26 = num("Customer %", 10, "D26", fmt="%.1f", step=0.5)

# # ── SECTION 3: Pricing ──
# st.markdown('<div class="section-header">SECTION 3 — Pricing</div>', unsafe_allow_html=True)
# col1, col2 = st.columns([4, 1])
# with col1:
#     st.markdown("Skonto / early-payment discount *(Baugruppenpreis zzgl. Skonto)*")
# with col2:
#     B31 = num("Skonto factor", 0.0, "B31", fmt="%.3f", step=0.01)

# col1, col2 = st.columns([4, 1])
# with col1:
#     st.markdown("Profit margin *(Baugruppenpreis)*")
# with col2:
#     B32 = num("Margin factor", 0.05, "B32", fmt="%.3f", step=0.005)

# # ── SECTION 4: Setup / Initial Costs ──
# st.markdown('<div class="section-header">SECTION 4 — Initial / Setup Costs</div>', unsafe_allow_html=True)

# setup_rows = [
#     ("SMD Assembly Setup per side (Valor, Rüstung, EFA, 3D-AOI)",    "C38", 150,  "D38", 90),
#     ("Setup Printer programme (Druckerprogramm)",                     "C39", 45,   "D39", 45),
#     ("Setup SMD per part (VPL, SiplacePro, Feeder, 3D-AOI)",          "C40", 6.75, "D40", 1.75),
#     ("Setup THT per part (Selektivprogramm)",                          "C41", 6,    "D41", 2.5),
# ]
# input_vals = {}
# for label, ck, cv, dk, dv in setup_rows:
#     col1, col2, col3 = st.columns([3, 1, 1])
#     with col1: st.markdown(label)
#     with col2: input_vals[ck] = num("Standard €", cv, ck, fmt="%.2f", step=0.25)
#     with col3: input_vals[dk] = num("Customer €", dv, dk, fmt="%.2f", step=0.25)
# C38, D38 = input_vals["C38"], input_vals["D38"]
# C39, D39 = input_vals["C39"], input_vals["D39"]
# C40, D40 = input_vals["C40"], input_vals["D40"]
# C41, D41 = input_vals["C41"], input_vals["D41"]

# col1, col2 = st.columns([4, 1])
# with col1: st.markdown("Setup LP Supplier *(Setup Cost LP Lieferant)*")
# with col2: D42 = num("Customer €", 0, "D42", fmt="%.2f", step=5.0)

# col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
# with col1: st.markdown("Stencil Top *(Pastensieb Top)*")
# with col2: C43 = num("Standard €", 150, "C43", fmt="%.2f", step=5.0)
# with col3: D43 = num("Customer €", 0,   "D43", fmt="%.2f", step=5.0)
# with col4: E43 = num("Surcharge factor", 0.2, "E43", fmt="%.2f")

# col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
# with col1: st.markdown("Stencil Bottom *(Pastensieb Bot)*")
# with col2: C44 = num("Standard €", 150, "C44", fmt="%.2f", step=5.0)
# with col3: D44 = num("Customer €", 0,   "D44", fmt="%.2f", step=5.0)
# with col4: E44 = num("Surcharge factor", 0.2, "E44", fmt="%.2f")

# col1, col2, col3 = st.columns([3, 1, 1])
# with col1: st.markdown("Order processing *(Auftragsbearbeitung)*")
# with col2: C45 = num("Standard €", 60, "C45", fmt="%.2f", step=5.0)
# with col3: D45 = num("Customer €", 60, "D45", fmt="%.2f", step=5.0)

# col1, col2 = st.columns([4, 1])
# with col1: st.markdown("STD lead time days *(STD und Tage)*")
# with col2: B52 = num_int("Days", 110, "B52")

# # ──────────────────────────────────────────────
# # CALCULATIONS — exact formulas from standard.xlsx
# # ──────────────────────────────────────────────

# # G2 = qty_std, H2 = qty_cust

# # Row 7
# G7 = C7 * E7 / 100
# H7 = D7 * E7 / 100

# # Row 8
# G8 = (C8 / E13 / 100) if E8 == 2 else 0
# H8 = (D8 / E13 / 100) if E8 == 2 else 0

# # Row 11 — =(((C9*E8)+(C10*E8)+(C11*E11))/G2)
# G11 = (((C9 * E8) + (C10 * E8) + (C11 * E11)) / qty_std) if qty_std else 0
# H11 = (((D9 * E8) + (D10 * E8) + (D11 * E11)) / qty_cust) if qty_cust else 0

# # Row 15 — =(((C14*E8)/G2)+((C15*E15)/100))
# G15 = (((C14 * E8) / qty_std) + ((C15 * E15) / 100)) if qty_std else 0
# H15 = (((D14 * E8) / qty_cust) + ((D15 * E15) / 100)) if qty_cust else 0

# # Row 16 — =C16*C18*E16/60
# G16 = C16 * C18 * E16 / 60
# H16 = D16 * D18 * E16 / 60

# # Row 17 — =F17*C18/60
# G17 = F17 * C18 / 60
# H17 = F17 * D18 / 60

# # Row 19 — =SUM(G7:G18)  (rows 9,10,12,13,14 are inputs only; G/H cols empty)
# G19 = G7 + G8 + G11 + G15 + G16 + G17
# H19 = H7 + H8 + H11 + H15 + H16 + H17

# # Row 25 — =(C25*C26/100)+C25
# G25 = (C25 * C26 / 100) + C25
# H25 = (D25 * D26 / 100) + D25

# # Row 27 — =C25*C26/100*G2
# C27 = C25 * C26 / 100 * qty_std
# D27 = D25 * D26 / 100 * qty_cust

# # Row 30 — =G25+G19
# C30 = G25 + G19
# D30 = H25 + H19

# # Row 31 — =C30*B31+C30
# C31 = C30 * B31 + C30
# D31 = D30 * B31 + D30

# # Row 32 — =C31*B32+C31
# G32 = C31 * B32 + C31
# H32 = D31 * B32 + D31

# # Row 33 — =G32*E13
# C33 = G32 * E13
# D33 = H32 * E13

# # Row 34 — =G32*G2 / =H2*H32
# G34 = G32 * qty_std
# H34 = qty_cust * H32

# # Row 35 — =G34-(C25*G2)
# G35 = G34 - (C25 * qty_std)
# H35 = H34 - (D25 * qty_cust)

# # Setup rows
# H38 = D38 * E8
# H39 = D39 * E8
# H40 = D40 * E11
# H41 = D41 * E15
# H42 = D42
# H43 = D43 + (D43 * E43)
# H44 = D44 + (D44 * E44)
# H45 = D45

# # Row 46 — =SUM(H38:H45)
# H46 = H38 + H39 + H40 + H41 + H42 + H43 + H44 + H45

# # Row 48 — =G34+H46   (project price for STD qty)
# H48 = G34 + H46
# # Row 49 — =H34+H46   (project price for CUST qty)
# H49 = H34 + H46

# # Row 52
# G52 = (G34 / B52 / 8) if B52 else 0
# H52 = (H34 / B52 / 8) if B52 else 0

# # ──────────────────────────────────────────────
# # RESULTS DISPLAY
# # ──────────────────────────────────────────────
# st.markdown("---")
# st.markdown("## 📊 Calculation Results")

# def fmt(v): return f"€ {v:,.4f}"

# col_s, col_c = st.columns(2)

# with col_s:
#     st.markdown(f'<div class="sub-header">🔵 Standard  (Qty: {qty_std})</div>', unsafe_allow_html=True)

#     st.markdown("**Labour Breakdown**")
#     st.markdown(f'<div class="result-box">SMD assembly: <b>{fmt(G7)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Soldering: <b>{fmt(G8)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">FIX + Feeder setup/unit: <b>{fmt(G11)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">THT cost/unit: <b>{fmt(G15)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Manual soldering: <b>{fmt(G16)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">QS: <b>{fmt(G17)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="total-box">Total Labour (no markup): <b>{fmt(G19)}</b></div>', unsafe_allow_html=True)

#     st.markdown("**Material**")
#     st.markdown(f'<div class="result-box">Material cost + markup/unit: <b>{fmt(G25)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Total material margin (order): <b>{fmt(C27)}</b></div>', unsafe_allow_html=True)

#     st.markdown("**Pricing**")
#     st.markdown(f'<div class="result-box">Assembly price incl. material: <b>{fmt(C30)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Price + skonto: <b>{fmt(C31)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Unit price incl. profit margin: <b>{fmt(G32)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Panel price (×{E13} circuits): <b>{fmt(C33)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="total-box">Total order value (LOHN+Mat): <b>{fmt(G34)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Order value LOHN only: <b>{fmt(G35)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Cost/hour (STD {B52}d×8h): <b>{fmt(G52)}</b></div>', unsafe_allow_html=True)

# with col_c:
#     st.markdown(f'<div class="sub-header">🟢 Customer  (Qty: {qty_cust})</div>', unsafe_allow_html=True)

#     st.markdown("**Labour Breakdown**")
#     st.markdown(f'<div class="result-box">SMD assembly: <b>{fmt(H7)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Soldering: <b>{fmt(H8)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">FIX + Feeder setup/unit: <b>{fmt(H11)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">THT cost/unit: <b>{fmt(H15)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Manual soldering: <b>{fmt(H16)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">QS: <b>{fmt(H17)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="total-box">Total Labour (no markup): <b>{fmt(H19)}</b></div>', unsafe_allow_html=True)

#     st.markdown("**Material**")
#     st.markdown(f'<div class="result-box">Material cost + markup/unit: <b>{fmt(H25)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Total material margin (order): <b>{fmt(D27)}</b></div>', unsafe_allow_html=True)

#     st.markdown("**Pricing**")
#     st.markdown(f'<div class="result-box">Assembly price incl. material: <b>{fmt(D30)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Price + skonto: <b>{fmt(D31)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Unit price incl. profit margin: <b>{fmt(H32)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Panel price (×{E13} circuits): <b>{fmt(D33)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="total-box">Total order value (LOHN+Mat): <b>{fmt(H34)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Order value LOHN only: <b>{fmt(H35)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Cost/hour (STD {B52}d×8h): <b>{fmt(H52)}</b></div>', unsafe_allow_html=True)

# # Setup summary
# st.markdown("---")
# st.markdown("### 🛠️ Initial / Setup Costs (Customer)")
# col1, col2, col3, col4 = st.columns(4)
# with col1:
#     st.markdown(f'<div class="result-box">SMD Setup: <b>{fmt(H38)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Printer: <b>{fmt(H39)}</b></div>', unsafe_allow_html=True)
# with col2:
#     st.markdown(f'<div class="result-box">SMD/part: <b>{fmt(H40)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">THT/part: <b>{fmt(H41)}</b></div>', unsafe_allow_html=True)
# with col3:
#     st.markdown(f'<div class="result-box">LP Supplier: <b>{fmt(H42)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Stencil Top: <b>{fmt(H43)}</b></div>', unsafe_allow_html=True)
# with col4:
#     st.markdown(f'<div class="result-box">Stencil Bot: <b>{fmt(H44)}</b></div>', unsafe_allow_html=True)
#     st.markdown(f'<div class="result-box">Order proc.: <b>{fmt(H45)}</b></div>', unsafe_allow_html=True)

# st.markdown(f'<div class="total-box">TOTAL Initial Costs: <b>{fmt(H46)}</b></div>', unsafe_allow_html=True)

# # Grand totals
# st.markdown("---")
# st.markdown("### 🏁 Project Summary")
# col1, col2 = st.columns(2)
# with col1:
#     st.markdown(f'<div class="grand-total">Project price for STD qty ({qty_std}) incl. setup:<br><b>{fmt(H48)}</b></div>', unsafe_allow_html=True)
# with col2:
#     st.markdown(f'<div class="grand-total">Project price for Customer qty ({qty_cust}) incl. setup:<br><b>{fmt(H49)}</b></div>', unsafe_allow_html=True)

# # ──────────────────────────────────────────────
# # EXCEL EXPORT — exact structure from standard.xlsx
# # ──────────────────────────────────────────────

# def build_excel():
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = f"{qty_cust} Stk. am {calc_date.strftime('%d.%m.%Y')}"

#     # Column widths (match original)
#     ws.column_dimensions['A'].width = 72
#     ws.column_dimensions['B'].width = 10
#     ws.column_dimensions['C'].width = 14
#     ws.column_dimensions['D'].width = 14
#     ws.column_dimensions['E'].width = 12
#     ws.column_dimensions['F'].width = 10
#     ws.column_dimensions['G'].width = 16
#     ws.column_dimensions['H'].width = 16

#     EUR = '#,##0.0000 €'
#     thin = Side(style="thin", color="AAAAAA")
#     bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     def c(row, col, value, bold=False, bg=None, color="000000", num_fmt=None, align="left", wrap=False):
#         cell = ws.cell(row=row, column=col, value=value)
#         cell.font = Font(name="Arial", size=10, bold=bold, color=color)
#         if bg:
#             cell.fill = PatternFill("solid", fgColor=bg)
#         cell.border = bdr
#         cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
#         if num_fmt:
#             cell.number_format = num_fmt
#         return cell

#     def formula(row, col, f, bold=False, bg="FFF2CC", num_fmt=EUR):
#         cell = ws.cell(row=row, column=col, value=f)
#         cell.font = Font(name="Arial", size=10, bold=bold, color="000000")
#         cell.fill = PatternFill("solid", fgColor=bg)
#         cell.border = bdr
#         cell.alignment = Alignment(horizontal="right", vertical="center")
#         cell.number_format = num_fmt
#         return cell

#     def inp(row, col, value, num_fmt=None, align="right"):
#         cell = ws.cell(row=row, column=col, value=value)
#         cell.font = Font(name="Arial", size=10, color="0000FF")  # blue = user input
#         cell.fill = PatternFill("solid", fgColor="E2EFDA")
#         cell.border = bdr
#         cell.alignment = Alignment(horizontal=align, vertical="center")
#         if num_fmt:
#             cell.number_format = num_fmt
#         return cell

#     # ── Row 1: headers
#     c(1, 7, "Teilsumme", bold=True, bg="1F4E79", color="FFFFFF", align="center")
#     c(1, 8, "Teilsumme", bold=True, bg="1F4E79", color="FFFFFF", align="center")

#     # ── Row 2: quantities  G2 / H2
#     inp(2, 7, qty_std,  num_fmt="0")
#     inp(2, 8, qty_cust, num_fmt="0")

#     # ── Row 4: Platine
#     c(4, 1, f"Platine:  {platine}", bold=True, wrap=True)

#     # ── Row 6: Section header
#     c(6, 1, "Baugruppe LOHN", bold=True, bg="2E75B6", color="FFFFFF")
#     c(6, 3, "amount/50",  bold=True, bg="2E75B6", color="FFFFFF", align="center")
#     c(6, 4, "amount/100", bold=True, bg="2E75B6", color="FFFFFF", align="center")
#     c(6, 5, "no. of components SMT", bold=True, bg="2E75B6", color="FFFFFF", align="center")
#     c(6, 7, "cost cal", bold=True, bg="2E75B6", color="FFFFFF", align="center")
#     c(6, 9, "Total cost", bold=True, bg="2E75B6", color="FFFFFF", align="center")

#     # ── Row 7: Bauteile SMD
#     ws.row_dimensions[7].height = 30
#     c(7, 1, "Bauteile SMD (Time needed to finish one assembly) /SMT components/ SMT", wrap=True)
#     c(7, 2, "Ct", align="center")
#     inp(7, 3, C7,  num_fmt="0.00")
#     inp(7, 4, D7,  num_fmt="0.00")
#     inp(7, 5, E7,  num_fmt="0")
#     formula(7, 7, "=C7*E7/100")
#     formula(7, 8, "=D7*E7/100")

#     # ── Row 8: Lötung
#     ws.row_dimensions[8].height = 30
#     c(8, 1, "Lötung 1- / 2-Seitig   / one-sided or two-sided", wrap=True)
#     c(8, 2, "Ct", align="center")
#     inp(8, 3, C8, num_fmt="0.00")
#     inp(8, 4, D8, num_fmt="0.00")
#     inp(8, 5, E8, num_fmt="0")
#     formula(8, 7, "=IF(E8=2,C8/E13/100,0)")
#     formula(8, 8, "=IF(E8=2,D8/E13/100,0)")

#     # ── Row 9: FIX Drucker
#     ws.row_dimensions[9].height = 24
#     c(9, 1, "FIX Drucker pro Seite ( 30,00 euro standard )   /Paste printer", wrap=True)
#     inp(9, 3, C9, num_fmt=EUR)
#     inp(9, 4, D9, num_fmt=EUR)

#     # ── Row 10: FIX Maschine
#     ws.row_dimensions[10].height = 24
#     c(10, 1, "FIX Maschine pro Seite ( 30,00 euro standard )  / fix Machine Cost", wrap=True)
#     inp(10, 3, C10, num_fmt=EUR)
#     inp(10, 4, D10, num_fmt=EUR)

#     # ── Row 11: Feeder
#     ws.row_dimensions[11].height = 36
#     c(11, 1, "Verschiedene SMD-Bauteile zum Rüsten (Feeder for all components/Teaching/Lager Material einrechten)  / types of components", wrap=True)
#     c(11, 2, "€", align="center")
#     inp(11, 3, C11, num_fmt=EUR)
#     inp(11, 4, D11, num_fmt=EUR)
#     inp(11, 5, E11, num_fmt="0")
#     formula(11, 7, "=(((C9*E8)+(C10*E8)+(C11*E11))/G2)")
#     formula(11, 8, "=(((D9*E8)+(D10*E8)+(D11*E11))/H2)")

#     # ── Row 12: THT types
#     ws.row_dimensions[12].height = 24
#     c(12, 1, "Verschiedene THT-Bauteile / diff type of tht components", wrap=True)
#     inp(12, 5, E12, num_fmt="0")

#     # ── Row 13: Einzelschaltungen
#     ws.row_dimensions[13].height = 24
#     c(13, 1, "Einzelschaltungen pro Nutzen / assemblies in one panel", wrap=True)
#     inp(13, 5, E13, num_fmt="0")

#     # ── Row 14: FIX Selektiv
#     ws.row_dimensions[14].height = 24
#     c(14, 1, "FIX Selektiv oder Welle ( 30,00 euro standard )  /selective solder", wrap=True)
#     c(14, 2, "€", align="center")
#     inp(14, 3, C14, num_fmt=EUR)
#     inp(14, 4, D14, num_fmt=EUR)

#     # ── Row 15: Bauteile THT
#     ws.row_dimensions[15].height = 30
#     c(15, 1, "Bauteile THT ( 35,00 euro standard )  / THT component cost /THT", wrap=True)
#     c(15, 2, "Ct", align="center")
#     inp(15, 3, C15, num_fmt="0.00")
#     inp(15, 4, D15, num_fmt="0.00")
#     inp(15, 5, E15, num_fmt="0")
#     formula(15, 7, "=(((C14*E8)/G2)+((C15*E15)/100))")
#     formula(15, 8, "=(((D14*E8)/H2)+((D15*E15)/100))")

#     # ── Row 16: Hand solder
#     ws.row_dimensions[16].height = 28
#     c(16, 1, "Einzellötstellen von Hand (0,08' -> 5 Sekunden) / hand solder", wrap=True)
#     inp(16, 3, C16, num_fmt="0")
#     inp(16, 4, D16, num_fmt="0")
#     inp(16, 5, E16, num_fmt="0.00")
#     formula(16, 7, "=C16*C18*E16/60")
#     formula(16, 8, "=D16*D18*E16/60")

#     # ── Row 17: QS
#     ws.row_dimensions[17].height = 24
#     c(17, 1, "QS / quality inspection", wrap=True)
#     inp(17, 6, F17, num_fmt="0.0")
#     formula(17, 7, "=F17*C18/60")
#     formula(17, 8, "=F17*D18/60")

#     # ── Row 18: Stundensatz
#     ws.row_dimensions[18].height = 24
#     c(18, 1, "Faktor Stundensatz  /hourly cost", wrap=True)
#     c(18, 2, "€", align="center")
#     inp(18, 3, C18, num_fmt=EUR)
#     inp(18, 4, D18, num_fmt=EUR)

#     # ── Row 19: Total Labour
#     ws.row_dimensions[19].height = 24
#     c(19, 1, "Summe Lohn ohne Aufschlag  /TOTAL", bold=True, bg="FFD966")
#     formula(19, 7, "=SUM(G7:G18)", bold=True, bg="FFD966")
#     formula(19, 8, "=SUM(H7:H18)", bold=True, bg="FFD966")

#     # ── Row 23: Material section
#     c(23, 1, "Baugruppe Material / assembly material cost", bold=True, bg="2E75B6", color="FFFFFF")

#     # ── Row 25: Materialkosten
#     ws.row_dimensions[25].height = 24
#     c(25, 1, "Materialkosten  /material cost", wrap=True)
#     c(25, 2, "€", align="center")
#     inp(25, 3, C25, num_fmt=EUR)
#     inp(25, 4, D25, num_fmt=EUR)
#     formula(25, 7, "=(C25*C26/100)+C25")
#     formula(25, 8, "=(D25*D26/100)+D25")

#     # ── Row 26: Aufschlag
#     ws.row_dimensions[26].height = 24
#     c(26, 1, "Aufschlag in %  /margin", wrap=True)
#     inp(26, 3, C26, num_fmt="0.0")
#     inp(26, 4, D26, num_fmt="0.0")

#     # ── Row 27: Materialaufschlag
#     ws.row_dimensions[27].height = 24
#     c(27, 1, "Materialaufschlag insgesamt  /only margin profit", wrap=True)
#     c(27, 2, "€", align="center")
#     formula(27, 3, "=C25*C26/100*G2")
#     formula(27, 4, "=D25*D26/100*H2")

#     # ── Row 30: Assembly price
#     ws.row_dimensions[30].height = 24
#     c(30, 1, "Baugruppepreis Lohn inkl. Material  /assembly cost inclusive materials", wrap=True)
#     formula(30, 3, "=G25+G19")
#     formula(30, 4, "=H25+H19")

#     # ── Row 31: Skonto
#     ws.row_dimensions[31].height = 24
#     c(31, 1, "Baugruppenpreis zzgl. Skonto  /advance margin", wrap=True)
#     inp(31, 2, B31, num_fmt="0.000")
#     formula(31, 3, "=C30*B31+C30")
#     formula(31, 4, "=D30*B31+D30")

#     # ── Row 32: Profit margin
#     ws.row_dimensions[32].height = 24
#     c(32, 1, "Baugruppenpreis  /assembly price", wrap=True)
#     inp(32, 2, B32, num_fmt="0.000")
#     formula(32, 7, "=C31*B32+C31")
#     formula(32, 8, "=D31*B32+D31")

#     # ── Row 33: Panel price
#     ws.row_dimensions[33].height = 24
#     c(33, 1, "Nutzenpreis inkl. Skonto und Gewinnmarge", wrap=True)
#     formula(33, 3, "=G32*E13")
#     formula(33, 4, "=H32*E13")

#     # ── Row 34: Total order
#     ws.row_dimensions[34].height = 24
#     c(34, 1, "Auftragswert LOHN + Material", bold=True, bg="FFD966")
#     formula(34, 7, "=G32*G2", bold=True, bg="FFD966")
#     formula(34, 8, "=H2*H32", bold=True, bg="FFD966")

#     # ── Row 35: LOHN only
#     ws.row_dimensions[35].height = 24
#     formula(35, 7, "=G34-(C25*G2)")
#     formula(35, 8, "=H34-(D25*H2)")

#     # ── Row 37: Standard label
#     c(37, 3, "Standard", bold=True, align="center")

#     # ── Row 38-45: Setup costs
#     setup_data = [
#         (38, "FIXKOSTEN Assembly SMD per SIDE (Valor, Rüstung erstellen, EFA, 3D-AOI)   /fix cost machine", C38, D38, "=D38*E8"),
#         (39, "Setup Cost Druckerprogramm erstellen /paste printer program cost",           C39, D39, "=D39*E8"),
#         (40, "Setup Cost SMD per Part (VPL Search, Gehäuse SiplacePro, Feeder-Setup, MagicKlick-3D-AOI-Optimierung) /set up cost per part", C40, D40, "=D40*E11"),
#         (41, "Setup Cost THT per Part (Selektivprogramm pro BT einstellen und optimieren) /set up cost for tht programming", C41, D41, "=D41*E15"),
#         (42, "Setup Cost LP Lieferant / PCB supplier",                                     0,   D42, "=D42"),
#         (43, "Pastensieb Top  / stencil top cost",                                          C43, D43, "=D43+(D43*E43)"),
#         (44, "Pastensieb Bot /stencil bottom cost",                                         C44, D44, "=D44+(D44*E44)"),
#         (45, "Auftragsbearbeitung (Fertigungsauftrag anlegen, Auftragsbestätiging erzeugen, Rechnungslabel /work processing cost", C45, D45, "=D45"),
#     ]
#     for row, label, std_v, cust_v, h_formula in setup_data:
#         ws.row_dimensions[row].height = 30
#         c(row, 1, label, wrap=True)
#         c(row, 2, "€", align="center")
#         inp(row, 3, std_v, num_fmt=EUR)
#         inp(row, 4, cust_v, num_fmt=EUR)
#         formula(row, 8, h_formula)

#     # Stencil factors
#     inp(43, 5, E43, num_fmt="0.0")
#     inp(44, 5, E44, num_fmt="0.0")

#     # ── Row 46: Total initial
#     ws.row_dimensions[46].height = 24
#     c(46, 1, "Gesamt Initialkosten  /total initial cost", bold=True, bg="FFD966")
#     formula(46, 8, "=SUM(H38:H45)", bold=True, bg="FFD966")

#     # ── Row 48: Project price std qty
#     ws.row_dimensions[48].height = 28
#     c(48, 1, f"Projektpreis inkl. Setupkosten für {qty_std} Baugruppen /project price for {qty_std}", bold=True, bg="FFF2CC", wrap=True)
#     formula(48, 8, "=G34+H46", bold=True, bg="FFF2CC")

#     # ── Row 49: Project price cust qty
#     ws.row_dimensions[49].height = 28
#     c(49, 1, f"Projektpreis Serienfertigung inkl. Setupkosten für {qty_cust} Baugruppen /project for {qty_cust}", bold=True, bg="FFD966", wrap=True)
#     formula(49, 8, "=H34+H46", bold=True, bg="FFD966")

#     # ── Row 52: STD days
#     ws.row_dimensions[52].height = 24
#     c(52, 1, "STD und Tage /", wrap=True)
#     inp(52, 2, B52, num_fmt="0")
#     c(52, 3, "€", align="center")
#     c(52, 4, "€", align="center")
#     formula(52, 7, "=G34/B52/8")
#     formula(52, 8, "=H34/B52/8")

#     # ── Footer
#     c(59, 1, f"Erstellt von {created_by}", bold=True)
#     c(60, 1, f"Datum: {calc_date.strftime('%d.%m.%Y')}")

#     # Freeze panes
#     ws.freeze_panes = "A7"

#     buf = io.BytesIO()
#     wb.save(buf)
#     buf.seek(0)
#     return buf

# st.markdown("---")
# st.markdown("### 📥 Export Excel")
# if st.button("⬇️ Generate Excel File", type="primary"):
#     buf = build_excel()
#     fname = f"Assembly_Cost_{qty_cust}pcs_{calc_date.strftime('%d%m%Y')}.xlsx"
#     st.download_button(
#         label="📎 Download Excel",
#         data=buf,
#         file_name=fname,
#         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     st.success(f"✅ Excel ready: {fname}")

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from datetime import date

st.set_page_config(page_title="PCB Assembly Cost Calculator", layout="wide")

st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; }
    .section-header {
        background: #1F4E79; color: white; padding: 8px 14px;
        border-radius: 6px; font-weight: bold; margin: 18px 0 8px 0;
    }
    .sub-header {
        background: #2E75B6; color: white; padding: 5px 12px;
        border-radius: 4px; font-weight: 600; margin: 12px 0 6px 0; font-size: 0.9rem;
    }
    .result-box {
        background: #EAF4FB; border: 1px solid #2E75B6;
        border-radius: 6px; padding: 10px 16px; margin: 4px 0; color: #1a1a1a;
    }
    .total-box {
        background: #2E75B6; border: 1px solid #1F4E79;
        border-radius: 6px; padding: 10px 16px; margin: 6px 0;
        font-weight: bold; color: #ffffff;
    }
    .grand-total {
        background: #1F4E79; color: white; border-radius: 8px;
        padding: 14px 20px; margin: 8px 0; font-size: 1.1rem; font-weight: bold;
    }
    .stNumberInput label, .stTextInput label { font-size: 0.85rem; }
</style>
""", unsafe_allow_html=True)

st.markdown("## 🔧 PCB Assembly Cost Calculator")
st.caption("Translated from: *Baugruppe LOHN* — exact formulas preserved from standard.xlsx")

# ──────────────────────────────────────────────
# SIDEBAR INPUTS
# ──────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ General Settings")
    qty_std  = st.number_input("Quantity Standard (G2)", value=50,  min_value=1, step=1)
    qty_cust = st.number_input("Quantity Customer (H2)", value=100, min_value=1, step=1)
    platine  = st.text_input("PCB / Platine name", value="PCB Assembly")
    created_by = st.text_input("Created by", value="Ashok Kumar")
    calc_date  = st.date_input("Date", value=date.today())

# ──────────────────────────────────────────────
# MAIN INPUTS — two column layout (Std | Cust)
# ──────────────────────────────────────────────

def num(label, val, key, fmt="%.2f", step=0.01, min_val=0.0):
    return st.number_input(label, value=float(val), key=key, format=fmt, step=step, min_value=min_val)

def num_int(label, val, key):
    return st.number_input(label, value=int(val), key=key, step=1, min_value=0)

# ── SECTION 1: SMD ──
st.markdown('<div class="section-header">SECTION 1 — Labour Assembly (LOHN)</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">SMD Components</div>', unsafe_allow_html=True)

col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.markdown("SMD Components — Time per assembly *(Bauteile SMD)*")
with col2:
    C7 = num("Standard [Ct]", 5.5, "C7")
with col3:
    D7 = num("Customer [Ct]", 4.0, "D7")

col1, col2 = st.columns([4, 1])
with col1:
    st.markdown("Number of SMD components")
with col2:
    E7 = num_int("Count", 120, "E7")

st.markdown('<div class="sub-header">Soldering</div>', unsafe_allow_html=True)
col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
with col1:
    st.markdown("Soldering 1-sided / 2-sided *(Lötung)*")
with col2:
    C8 = num("Standard [Ct]", 150, "C8", fmt="%.1f", step=1.0)
with col3:
    D8 = num("Customer [Ct]", 150, "D8", fmt="%.1f", step=1.0)
with col4:
    E8 = st.number_input("Sides (1 or 2)", value=2, min_value=1, max_value=2, step=1, key="E8")

st.markdown('<div class="sub-header">FIX Costs per Side</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.markdown("FIX Printer per side *(Drucker)*")
with col2:
    C9 = num("Standard €", 30, "C9")
with col3:
    D9 = num("Customer €", 30, "D9")

col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.markdown("FIX Machine per side *(Maschine)*")
with col2:
    C10 = num("Standard €", 30, "C10")
with col3:
    D10 = num("Customer €", 30, "D10")

st.markdown('<div class="sub-header">SMD Feeder / Setup Types</div>', unsafe_allow_html=True)
col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
with col1:
    st.markdown("SMD feeder types *(Verschiedene SMD-Bauteile)*")
with col2:
    C11 = num("Standard €/type", 6.5, "C11")
with col3:
    D11 = num("Customer €/type", 6.5, "D11")
with col4:
    E11 = num_int("No. of types", 33, "E11")

st.markdown('<div class="sub-header">THT Components</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.markdown("THT component types *(Verschiedene THT-Bauteile)*")
with col2:
    E12 = num_int("THT types", 0, "E12")

col1, col2 = st.columns([4, 1])
with col1:
    st.markdown("Circuits per panel *(Einzelschaltungen pro Nutzen)*")
with col2:
    E13 = num_int("Count", 5, "E13")

col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.markdown("FIX Selective / Wave solder *(Selektiv oder Welle)*")
with col2:
    C14 = num("Standard €", 30, "C14")
with col3:
    D14 = num("Customer €", 30, "D14")

col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
with col1:
    st.markdown("THT component cost *(Bauteile THT)*")
with col2:
    C15 = num("Standard [Ct]", 55, "C15", fmt="%.2f", step=0.5)
with col3:
    D15 = num("Customer [Ct]", 45, "D15", fmt="%.2f", step=0.5)
with col4:
    E15 = num_int("THT qty", 7, "E15")

st.markdown('<div class="sub-header">Manual Soldering & QS</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.markdown("Manual solder joints *(Einzellötstellen von Hand)*")
with col2:
    C16 = num("Standard count", 0, "C16", fmt="%.0f", step=1.0)
with col3:
    D16 = num("Customer count", 0, "D16", fmt="%.0f", step=1.0)

col1, col2 = st.columns([4, 1])
with col1:
    st.markdown("Time per joint [min]")
with col2:
    E16 = num("Min/joint", 0.0, "E16")

col1, col2 = st.columns([4, 1])
with col1:
    st.markdown("QS / Quality inspection time [min]")
with col2:
    F17 = num("QS min", 1.8, "F17")

col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.markdown("Hourly rate *(Faktor Stundensatz)*")
with col2:
    C18 = num("Standard €/h", 60, "C18", fmt="%.2f", step=1.0)
with col3:
    D18 = num("Customer €/h", 60, "D18", fmt="%.2f", step=1.0)

# ── SECTION 2: Material ──
st.markdown('<div class="section-header">SECTION 2 — Material</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.markdown("Material cost per unit *(Materialkosten)*")
with col2:
    C25 = num("Standard €", 735, "C25", fmt="%.2f", step=1.0)
with col3:
    D25 = num("Customer €", 700, "D25", fmt="%.2f", step=1.0)

col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.markdown("Material markup % *(Aufschlag in %)*")
with col2:
    C26 = num("Standard %", 7, "C26", fmt="%.1f", step=0.5)
with col3:
    D26 = num("Customer %", 10, "D26", fmt="%.1f", step=0.5)

# ── SECTION 3: Pricing ──
st.markdown('<div class="section-header">SECTION 3 — Pricing</div>', unsafe_allow_html=True)
col1, col2 = st.columns([4, 1])
with col1:
    st.markdown("Skonto / early-payment discount *(Baugruppenpreis zzgl. Skonto)*")
with col2:
    B31 = num("Skonto factor", 0.0, "B31", fmt="%.3f", step=0.01)

col1, col2 = st.columns([4, 1])
with col1:
    st.markdown("Profit margin *(Baugruppenpreis)*")
with col2:
    B32 = num("Margin factor", 0.05, "B32", fmt="%.3f", step=0.005)

# ── SECTION 4: Setup / Initial Costs ──
st.markdown('<div class="section-header">SECTION 4 — Initial / Setup Costs</div>', unsafe_allow_html=True)

setup_rows = [
    ("SMD Assembly Setup per side (Valor, Rüstung, EFA, 3D-AOI)",    "C38", 150,  "D38", 90),
    ("Setup Printer programme (Druckerprogramm)",                     "C39", 45,   "D39", 45),
    ("Setup SMD per part (VPL, SiplacePro, Feeder, 3D-AOI)",          "C40", 6.75, "D40", 1.75),
    ("Setup THT per part (Selektivprogramm)",                          "C41", 6,    "D41", 2.5),
]
input_vals = {}
for label, ck, cv, dk, dv in setup_rows:
    col1, col2, col3 = st.columns([3, 1, 1])
    with col1: st.markdown(label)
    with col2: input_vals[ck] = num("Standard €", cv, ck, fmt="%.2f", step=0.25)
    with col3: input_vals[dk] = num("Customer €", dv, dk, fmt="%.2f", step=0.25)
C38, D38 = input_vals["C38"], input_vals["D38"]
C39, D39 = input_vals["C39"], input_vals["D39"]
C40, D40 = input_vals["C40"], input_vals["D40"]
C41, D41 = input_vals["C41"], input_vals["D41"]

col1, col2 = st.columns([4, 1])
with col1: st.markdown("Setup LP Supplier *(Setup Cost LP Lieferant)*")
with col2: D42 = num("Customer €", 0, "D42", fmt="%.2f", step=5.0)

col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
with col1: st.markdown("Stencil Top *(Pastensieb Top)*")
with col2: C43 = num("Standard €", 150, "C43", fmt="%.2f", step=5.0)
with col3: D43 = num("Customer €", 0,   "D43", fmt="%.2f", step=5.0)
with col4: E43 = num("Surcharge factor", 0.2, "E43", fmt="%.2f")

col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
with col1: st.markdown("Stencil Bottom *(Pastensieb Bot)*")
with col2: C44 = num("Standard €", 150, "C44", fmt="%.2f", step=5.0)
with col3: D44 = num("Customer €", 0,   "D44", fmt="%.2f", step=5.0)
with col4: E44 = num("Surcharge factor", 0.2, "E44", fmt="%.2f")

col1, col2, col3 = st.columns([3, 1, 1])
with col1: st.markdown("Order processing *(Auftragsbearbeitung)*")
with col2: C45 = num("Standard €", 60, "C45", fmt="%.2f", step=5.0)
with col3: D45 = num("Customer €", 60, "D45", fmt="%.2f", step=5.0)

col1, col2 = st.columns([4, 1])
with col1: st.markdown("STD lead time days *(STD und Tage)*")
with col2: B52 = num_int("Days", 110, "B52")

# ──────────────────────────────────────────────
# CALCULATIONS — exact formulas from standard.xlsx
# ──────────────────────────────────────────────

# G2 = qty_std, H2 = qty_cust

# Row 7
G7 = C7 * E7 / 100
H7 = D7 * E7 / 100

# Row 8
G8 = (C8 / E13 / 100) if E8 == 2 else 0
H8 = (D8 / E13 / 100) if E8 == 2 else 0

# Row 11 — =(((C9*E8)+(C10*E8)+(C11*E11))/G2)
G11 = (((C9 * E8) + (C10 * E8) + (C11 * E11)) / qty_std) if qty_std else 0
H11 = (((D9 * E8) + (D10 * E8) + (D11 * E11)) / qty_cust) if qty_cust else 0

# Row 15 — =(((C14*E8)/G2)+((C15*E15)/100))
G15 = (((C14 * E8) / qty_std) + ((C15 * E15) / 100)) if qty_std else 0
H15 = (((D14 * E8) / qty_cust) + ((D15 * E15) / 100)) if qty_cust else 0

# Row 16 — =C16*C18*E16/60
G16 = C16 * C18 * E16 / 60
H16 = D16 * D18 * E16 / 60

# Row 17 — =F17*C18/60
G17 = F17 * C18 / 60
H17 = F17 * D18 / 60

# Row 19 — =SUM(G7:G18)  (rows 9,10,12,13,14 are inputs only; G/H cols empty)
G19 = G7 + G8 + G11 + G15 + G16 + G17
H19 = H7 + H8 + H11 + H15 + H16 + H17

# Row 25 — =(C25*C26/100)+C25
G25 = (C25 * C26 / 100) + C25
H25 = (D25 * D26 / 100) + D25

# Row 27 — =C25*C26/100*G2
C27 = C25 * C26 / 100 * qty_std
D27 = D25 * D26 / 100 * qty_cust

# Row 30 — =G25+G19
C30 = G25 + G19
D30 = H25 + H19

# Row 31 — =C30*B31+C30
C31 = C30 * B31 + C30
D31 = D30 * B31 + D30

# Row 32 — =C31*B32+C31
G32 = C31 * B32 + C31
H32 = D31 * B32 + D31

# Row 33 — =G32*E13
C33 = G32 * E13
D33 = H32 * E13

# Row 34 — =G32*G2 / =H2*H32
G34 = G32 * qty_std
H34 = qty_cust * H32

# Row 35 — =G34-(C25*G2)
G35 = G34 - (C25 * qty_std)
H35 = H34 - (D25 * qty_cust)

# Setup rows
H38 = D38 * E8
H39 = D39 * E8
H40 = D40 * E11
H41 = D41 * E15
H42 = D42
H43 = D43 + (D43 * E43)
H44 = D44 + (D44 * E44)
H45 = D45

# Row 46 — =SUM(H38:H45)
H46 = H38 + H39 + H40 + H41 + H42 + H43 + H44 + H45

# Row 48 — =G34+H46   (project price for STD qty)
H48 = G34 + H46
# Row 49 — =H34+H46   (project price for CUST qty)
H49 = H34 + H46

# Row 52
G52 = (G34 / B52 / 8) if B52 else 0
H52 = (H34 / B52 / 8) if B52 else 0

# ──────────────────────────────────────────────
# RESULTS DISPLAY
# ──────────────────────────────────────────────
st.markdown("---")
st.markdown("## 📊 Calculation Results")

def fmt(v): return f"€ {v:,.4f}"

col_s, col_c = st.columns(2)

with col_s:
    st.markdown(f'<div class="sub-header">🔵 Standard  (Qty: {qty_std})</div>', unsafe_allow_html=True)

    st.markdown("**Labour Breakdown**")
    st.markdown(f'<div class="result-box">SMD assembly: <b>{fmt(G7)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Soldering: <b>{fmt(G8)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">FIX + Feeder setup/unit: <b>{fmt(G11)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">THT cost/unit: <b>{fmt(G15)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Manual soldering: <b>{fmt(G16)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">QS: <b>{fmt(G17)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="total-box">Total Labour (no markup): <b>{fmt(G19)}</b></div>', unsafe_allow_html=True)

    st.markdown("**Material**")
    st.markdown(f'<div class="result-box">Material cost + markup/unit: <b>{fmt(G25)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Total material margin (order): <b>{fmt(C27)}</b></div>', unsafe_allow_html=True)

    st.markdown("**Pricing**")
    st.markdown(f'<div class="result-box">Assembly price incl. material: <b>{fmt(C30)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Price + skonto: <b>{fmt(C31)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Unit price incl. profit margin: <b>{fmt(G32)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Panel price (×{E13} circuits): <b>{fmt(C33)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="total-box">Total order value (LOHN+Mat): <b>{fmt(G34)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Order value LOHN only: <b>{fmt(G35)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Cost/hour (STD {B52}d×8h): <b>{fmt(G52)}</b></div>', unsafe_allow_html=True)

with col_c:
    st.markdown(f'<div class="sub-header">🟢 Customer  (Qty: {qty_cust})</div>', unsafe_allow_html=True)

    st.markdown("**Labour Breakdown**")
    st.markdown(f'<div class="result-box">SMD assembly: <b>{fmt(H7)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Soldering: <b>{fmt(H8)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">FIX + Feeder setup/unit: <b>{fmt(H11)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">THT cost/unit: <b>{fmt(H15)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Manual soldering: <b>{fmt(H16)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">QS: <b>{fmt(H17)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="total-box">Total Labour (no markup): <b>{fmt(H19)}</b></div>', unsafe_allow_html=True)

    st.markdown("**Material**")
    st.markdown(f'<div class="result-box">Material cost + markup/unit: <b>{fmt(H25)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Total material margin (order): <b>{fmt(D27)}</b></div>', unsafe_allow_html=True)

    st.markdown("**Pricing**")
    st.markdown(f'<div class="result-box">Assembly price incl. material: <b>{fmt(D30)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Price + skonto: <b>{fmt(D31)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Unit price incl. profit margin: <b>{fmt(H32)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Panel price (×{E13} circuits): <b>{fmt(D33)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="total-box">Total order value (LOHN+Mat): <b>{fmt(H34)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Order value LOHN only: <b>{fmt(H35)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Cost/hour (STD {B52}d×8h): <b>{fmt(H52)}</b></div>', unsafe_allow_html=True)

# Setup summary
st.markdown("---")
st.markdown("### 🛠️ Initial / Setup Costs (Customer)")
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.markdown(f'<div class="result-box">SMD Setup: <b>{fmt(H38)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Printer: <b>{fmt(H39)}</b></div>', unsafe_allow_html=True)
with col2:
    st.markdown(f'<div class="result-box">SMD/part: <b>{fmt(H40)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">THT/part: <b>{fmt(H41)}</b></div>', unsafe_allow_html=True)
with col3:
    st.markdown(f'<div class="result-box">LP Supplier: <b>{fmt(H42)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Stencil Top: <b>{fmt(H43)}</b></div>', unsafe_allow_html=True)
with col4:
    st.markdown(f'<div class="result-box">Stencil Bot: <b>{fmt(H44)}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">Order proc.: <b>{fmt(H45)}</b></div>', unsafe_allow_html=True)

st.markdown(f'<div class="total-box">TOTAL Initial Costs: <b>{fmt(H46)}</b></div>', unsafe_allow_html=True)

# Grand totals
st.markdown("---")
st.markdown("### 🏁 Project Summary")
col1, col2 = st.columns(2)
with col1:
    st.markdown(f'<div class="grand-total">Project price for STD qty ({qty_std}) incl. setup:<br><b>{fmt(H48)}</b></div>', unsafe_allow_html=True)
with col2:
    st.markdown(f'<div class="grand-total">Project price for Customer qty ({qty_cust}) incl. setup:<br><b>{fmt(H49)}</b></div>', unsafe_allow_html=True)

# ──────────────────────────────────────────────
# EXCEL EXPORT — exact structure from standard.xlsx
# ──────────────────────────────────────────────

def build_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{qty_cust} Stk. am {calc_date.strftime('%d.%m.%Y')}"

    # Column widths (match original)
    ws.column_dimensions['A'].width = 72
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16

    EUR = '#,##0.0000 €'
    thin = Side(style="thin", color="AAAAAA")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def c(row, col, value, bold=False, bg=None, color="000000", num_fmt=None, align="left", wrap=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", size=10, bold=bold, color=color)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = bdr
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    def formula(row, col, f, bold=False, bg="FFF2CC", num_fmt=EUR):
        cell = ws.cell(row=row, column=col, value=f)
        cell.font = Font(name="Arial", size=10, bold=bold, color="000000")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = bdr
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = num_fmt
        return cell

    def inp(row, col, value, num_fmt=None, align="right"):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", size=10, color="0000FF")  # blue = user input
        cell.fill = PatternFill("solid", fgColor="E2EFDA")
        cell.border = bdr
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    # ── Row 1: headers
    c(1, 7, "Teilsumme", bold=True, bg="1F4E79", color="FFFFFF", align="center")
    c(1, 8, "Teilsumme", bold=True, bg="1F4E79", color="FFFFFF", align="center")

    # ── Row 2: quantities  G2 / H2
    inp(2, 7, qty_std,  num_fmt="0")
    inp(2, 8, qty_cust, num_fmt="0")

    # ── Row 4: Platine
    c(4, 1, f"Platine:  {platine}", bold=True, wrap=True)

    # ── Row 6: Section header
    c(6, 1, "Baugruppe LOHN", bold=True, bg="2E75B6", color="FFFFFF")
    c(6, 3, "amount/50",  bold=True, bg="2E75B6", color="FFFFFF", align="center")
    c(6, 4, "amount/100", bold=True, bg="2E75B6", color="FFFFFF", align="center")
    c(6, 5, "no. of components SMT", bold=True, bg="2E75B6", color="FFFFFF", align="center")
    c(6, 7, "cost cal", bold=True, bg="2E75B6", color="FFFFFF", align="center")
    c(6, 9, "Total cost", bold=True, bg="2E75B6", color="FFFFFF", align="center")

    # ── Row 7: Bauteile SMD
    ws.row_dimensions[7].height = 30
    c(7, 1, "Bauteile SMD (Time needed to finish one assembly) /SMT components/ SMT", wrap=True)
    c(7, 2, "Ct", align="center")
    inp(7, 3, C7,  num_fmt="0.00")
    inp(7, 4, D7,  num_fmt="0.00")
    inp(7, 5, E7,  num_fmt="0")
    formula(7, 7, "=C7*E7/100")
    formula(7, 8, "=D7*E7/100")

    # ── Row 8: Lötung
    ws.row_dimensions[8].height = 30
    c(8, 1, "Lötung 1- / 2-Seitig   / one-sided or two-sided", wrap=True)
    c(8, 2, "Ct", align="center")
    inp(8, 3, C8, num_fmt="0.00")
    inp(8, 4, D8, num_fmt="0.00")
    inp(8, 5, E8, num_fmt="0")
    formula(8, 7, "=IF(E8=2,C8/E13/100,0)")
    formula(8, 8, "=IF(E8=2,D8/E13/100,0)")

    # ── Row 9: FIX Drucker
    ws.row_dimensions[9].height = 24
    c(9, 1, "FIX Drucker pro Seite ( 30,00 euro standard )   /Paste printer", wrap=True)
    inp(9, 3, C9, num_fmt=EUR)
    inp(9, 4, D9, num_fmt=EUR)

    # ── Row 10: FIX Maschine
    ws.row_dimensions[10].height = 24
    c(10, 1, "FIX Maschine pro Seite ( 30,00 euro standard )  / fix Machine Cost", wrap=True)
    inp(10, 3, C10, num_fmt=EUR)
    inp(10, 4, D10, num_fmt=EUR)

    # ── Row 11: Feeder
    ws.row_dimensions[11].height = 36
    c(11, 1, "Verschiedene SMD-Bauteile zum Rüsten (Feeder for all components/Teaching/Lager Material einrechten)  / types of components", wrap=True)
    c(11, 2, "€", align="center")
    inp(11, 3, C11, num_fmt=EUR)
    inp(11, 4, D11, num_fmt=EUR)
    inp(11, 5, E11, num_fmt="0")
    formula(11, 7, "=(((C9*E8)+(C10*E8)+(C11*E11))/G2)")
    formula(11, 8, "=(((D9*E8)+(D10*E8)+(D11*E11))/H2)")

    # ── Row 12: THT types
    ws.row_dimensions[12].height = 24
    c(12, 1, "Verschiedene THT-Bauteile / diff type of tht components", wrap=True)
    inp(12, 5, E12, num_fmt="0")

    # ── Row 13: Einzelschaltungen
    ws.row_dimensions[13].height = 24
    c(13, 1, "Einzelschaltungen pro Nutzen / assemblies in one panel", wrap=True)
    inp(13, 5, E13, num_fmt="0")

    # ── Row 14: FIX Selektiv
    ws.row_dimensions[14].height = 24
    c(14, 1, "FIX Selektiv oder Welle ( 30,00 euro standard )  /selective solder", wrap=True)
    c(14, 2, "€", align="center")
    inp(14, 3, C14, num_fmt=EUR)
    inp(14, 4, D14, num_fmt=EUR)

    # ── Row 15: Bauteile THT
    ws.row_dimensions[15].height = 30
    c(15, 1, "Bauteile THT ( 35,00 euro standard )  / THT component cost /THT", wrap=True)
    c(15, 2, "Ct", align="center")
    inp(15, 3, C15, num_fmt="0.00")
    inp(15, 4, D15, num_fmt="0.00")
    inp(15, 5, E15, num_fmt="0")
    formula(15, 7, "=(((C14*E8)/G2)+((C15*E15)/100))")
    formula(15, 8, "=(((D14*E8)/H2)+((D15*E15)/100))")

    # ── Row 16: Hand solder
    ws.row_dimensions[16].height = 28
    c(16, 1, "Einzellötstellen von Hand (0,08' -> 5 Sekunden) / hand solder", wrap=True)
    inp(16, 3, C16, num_fmt="0")
    inp(16, 4, D16, num_fmt="0")
    inp(16, 5, E16, num_fmt="0.00")
    formula(16, 7, "=C16*C18*E16/60")
    formula(16, 8, "=D16*D18*E16/60")

    # ── Row 17: QS
    ws.row_dimensions[17].height = 24
    c(17, 1, "QS / quality inspection", wrap=True)
    inp(17, 6, F17, num_fmt="0.0")
    formula(17, 7, "=F17*C18/60")
    formula(17, 8, "=F17*D18/60")

    # ── Row 18: Stundensatz
    ws.row_dimensions[18].height = 24
    c(18, 1, "Faktor Stundensatz  /hourly cost", wrap=True)
    c(18, 2, "€", align="center")
    inp(18, 3, C18, num_fmt=EUR)
    inp(18, 4, D18, num_fmt=EUR)

    # ── Row 19: Total Labour
    ws.row_dimensions[19].height = 24
    c(19, 1, "Summe Lohn ohne Aufschlag  /TOTAL", bold=True, bg="FFD966")
    formula(19, 7, "=SUM(G7:G18)", bold=True, bg="FFD966")
    formula(19, 8, "=SUM(H7:H18)", bold=True, bg="FFD966")

    # ── Row 23: Material section
    c(23, 1, "Baugruppe Material / assembly material cost", bold=True, bg="2E75B6", color="FFFFFF")

    # ── Row 25: Materialkosten
    ws.row_dimensions[25].height = 24
    c(25, 1, "Materialkosten  /material cost", wrap=True)
    c(25, 2, "€", align="center")
    inp(25, 3, C25, num_fmt=EUR)
    inp(25, 4, D25, num_fmt=EUR)
    formula(25, 7, "=(C25*C26/100)+C25")
    formula(25, 8, "=(D25*D26/100)+D25")

    # ── Row 26: Aufschlag
    ws.row_dimensions[26].height = 24
    c(26, 1, "Aufschlag in %  /margin", wrap=True)
    inp(26, 3, C26, num_fmt="0.0")
    inp(26, 4, D26, num_fmt="0.0")

    # ── Row 27: Materialaufschlag
    ws.row_dimensions[27].height = 24
    c(27, 1, "Materialaufschlag insgesamt  /only margin profit", wrap=True)
    c(27, 2, "€", align="center")
    formula(27, 3, "=C25*C26/100*G2")
    formula(27, 4, "=D25*D26/100*H2")

    # ── Row 30: Assembly price
    ws.row_dimensions[30].height = 24
    c(30, 1, "Baugruppepreis Lohn inkl. Material  /assembly cost inclusive materials", wrap=True)
    formula(30, 3, "=G25+G19")
    formula(30, 4, "=H25+H19")

    # ── Row 31: Skonto
    ws.row_dimensions[31].height = 24
    c(31, 1, "Baugruppenpreis zzgl. Skonto  /advance margin", wrap=True)
    inp(31, 2, B31, num_fmt="0.000")
    formula(31, 3, "=C30*B31+C30")
    formula(31, 4, "=D30*B31+D30")

    # ── Row 32: Profit margin
    ws.row_dimensions[32].height = 24
    c(32, 1, "Baugruppenpreis  /assembly price", wrap=True)
    inp(32, 2, B32, num_fmt="0.000")
    formula(32, 7, "=C31*B32+C31")
    formula(32, 8, "=D31*B32+D31")

    # ── Row 33: Panel price
    ws.row_dimensions[33].height = 24
    c(33, 1, "Nutzenpreis inkl. Skonto und Gewinnmarge", wrap=True)
    formula(33, 3, "=G32*E13")
    formula(33, 4, "=H32*E13")

    # ── Row 34: Total order
    ws.row_dimensions[34].height = 24
    c(34, 1, "Auftragswert LOHN + Material", bold=True, bg="FFD966")
    formula(34, 7, "=G32*G2", bold=True, bg="FFD966")
    formula(34, 8, "=H2*H32", bold=True, bg="FFD966")

    # ── Row 35: LOHN only
    ws.row_dimensions[35].height = 24
    formula(35, 7, "=G34-(C25*G2)")
    formula(35, 8, "=H34-(D25*H2)")

    # ── Row 37: Standard label
    c(37, 3, "Standard", bold=True, align="center")

    # ── Row 38-45: Setup costs
    setup_data = [
        (38, "FIXKOSTEN Assembly SMD per SIDE (Valor, Rüstung erstellen, EFA, 3D-AOI)   /fix cost machine", C38, D38, "=D38*E8"),
        (39, "Setup Cost Druckerprogramm erstellen /paste printer program cost",           C39, D39, "=D39*E8"),
        (40, "Setup Cost SMD per Part (VPL Search, Gehäuse SiplacePro, Feeder-Setup, MagicKlick-3D-AOI-Optimierung) /set up cost per part", C40, D40, "=D40*E11"),
        (41, "Setup Cost THT per Part (Selektivprogramm pro BT einstellen und optimieren) /set up cost for tht programming", C41, D41, "=D41*E15"),
        (42, "Setup Cost LP Lieferant / PCB supplier",                                     0,   D42, "=D42"),
        (43, "Pastensieb Top  / stencil top cost",                                          C43, D43, "=D43+(D43*E43)"),
        (44, "Pastensieb Bot /stencil bottom cost",                                         C44, D44, "=D44+(D44*E44)"),
        (45, "Auftragsbearbeitung (Fertigungsauftrag anlegen, Auftragsbestätiging erzeugen, Rechnungslabel /work processing cost", C45, D45, "=D45"),
    ]
    for row, label, std_v, cust_v, h_formula in setup_data:
        ws.row_dimensions[row].height = 30
        c(row, 1, label, wrap=True)
        c(row, 2, "€", align="center")
        inp(row, 3, std_v, num_fmt=EUR)
        inp(row, 4, cust_v, num_fmt=EUR)
        formula(row, 8, h_formula)

    # Stencil factors
    inp(43, 5, E43, num_fmt="0.0")
    inp(44, 5, E44, num_fmt="0.0")

    # ── Row 46: Total initial
    ws.row_dimensions[46].height = 24
    c(46, 1, "Gesamt Initialkosten  /total initial cost", bold=True, bg="FFD966")
    formula(46, 8, "=SUM(H38:H45)", bold=True, bg="FFD966")

    # ── Row 48: Project price std qty
    ws.row_dimensions[48].height = 28
    c(48, 1, f"Projektpreis inkl. Setupkosten für {qty_std} Baugruppen /project price for {qty_std}", bold=True, bg="FFF2CC", wrap=True)
    formula(48, 8, "=G34+H46", bold=True, bg="FFF2CC")

    # ── Row 49: Project price cust qty
    ws.row_dimensions[49].height = 28
    c(49, 1, f"Projektpreis Serienfertigung inkl. Setupkosten für {qty_cust} Baugruppen /project for {qty_cust}", bold=True, bg="FFD966", wrap=True)
    formula(49, 8, "=H34+H46", bold=True, bg="FFD966")

    # ── Row 52: STD days
    ws.row_dimensions[52].height = 24
    c(52, 1, "STD und Tage /", wrap=True)
    inp(52, 2, B52, num_fmt="0")
    c(52, 3, "€", align="center")
    c(52, 4, "€", align="center")
    formula(52, 7, "=G34/B52/8")
    formula(52, 8, "=H34/B52/8")

    # ── Footer
    c(59, 1, f"Erstellt von {created_by}", bold=True)
    c(60, 1, f"Datum: {calc_date.strftime('%d.%m.%Y')}")

    # Freeze panes
    ws.freeze_panes = "A7"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

st.markdown("---")
st.markdown("### 📥 Export Excel")
if st.button("⬇️ Generate Excel File", type="primary"):
    buf = build_excel()
    fname = f"Assembly_Cost_{qty_cust}pcs_{calc_date.strftime('%d%m%Y')}.xlsx"
    st.download_button(
        label="📎 Download Excel",
        data=buf,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.success(f"✅ Excel ready: {fname}")
