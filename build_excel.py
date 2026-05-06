"""Excel builder – returns .xlsx as bytes (no file I/O needed for Streamlit)."""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Styles ────────────────────────────────────────────────────────────────────

def _font(bold=False, size=10, color="000000", name="Arial"):
    return Font(name=name, size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL  = _fill("1F4E79");  SECTION_FILL = _fill("2E75B6")
INPUT_FILL   = _fill("E2EFDA");  LABEL_FILL   = _fill("D6DCE4")
RESULT_FILL  = _fill("FFF2CC");  TOTAL_FILL   = _fill("FFD966")
WHITE_FILL   = _fill("FFFFFF")

HEADER_FONT  = _font(bold=True,  size=11, color="FFFFFF")
SECTION_FONT = _font(bold=True,  size=10, color="FFFFFF")
LABEL_FONT   = _font(bold=False, size=10)
INPUT_FONT   = _font(bold=False, size=10, color="0000FF")
FORMULA_FONT = _font(bold=False, size=10, color="000000")
TOTAL_FONT   = _font(bold=True,  size=10)


def _style(cell, font=None, fill=None, align="left", border=True, num_fmt=None):
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if border: cell.border = _border()
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if num_fmt: cell.number_format = num_fmt


def build_excel_bytes(p: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{p['quantity']} pcs – {p['today']}"

    for col, w in [('A', 58), ('B', 10), ('C', 14), ('D', 14),
                   ('E', 12), ('F', 12), ('G', 16), ('H', 16)]:
        ws.column_dimensions[col].width = w

    EUR = '#,##0.00 €'

    def W(row, col, val, font=None, fill=None, align="left", num_fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        _style(c, font=font, fill=fill, align=align, num_fmt=num_fmt)
        return c

    # ── Title ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value=f"PCB ASSEMBLY COST SHEET  |  Qty: {p['quantity']}  |  Date: {p['today']}")
    _style(c, font=HEADER_FONT, fill=HEADER_FILL, align="center")

    # Legend
    ws.row_dimensions[2].height = 18
    W(2, 1, "Legend:", font=_font(bold=True, size=9))
    W(2, 3, "Standard",         font=_font(bold=True, size=9, color="FFFFFF"), fill=SECTION_FILL,      align="center")
    W(2, 4, "Customer",         font=_font(bold=True, size=9, color="FFFFFF"), fill=_fill("375623"),   align="center")
    W(2, 7, "Subtotal Standard",font=_font(bold=True, size=9, color="FFFFFF"), fill=SECTION_FILL,      align="center")
    W(2, 8, "Subtotal Customer",font=_font(bold=True, size=9, color="FFFFFF"), fill=_fill("375623"),   align="center")

    # Quantity
    ws.row_dimensions[3].height = 18
    W(3, 1, "PCB:", font=_font(bold=True))
    for col in [7, 8]:
        c = ws.cell(row=3, column=col, value=p['quantity'])
        _style(c, font=INPUT_FONT, fill=INPUT_FILL, align="center")

    # ── SECTION 1: Labour ────────────────────────────────────────────────────
    ROW = 5
    ws.row_dimensions[ROW].height = 20
    ws.merge_cells(f"A{ROW}:H{ROW}")
    _style(ws.cell(row=ROW, column=1, value="SECTION 1 – Labour Assembly (LOHN)"),
           font=SECTION_FONT, fill=SECTION_FILL)

    headers = ["Description", "Unit", "Standard", "Customer", "Qty / Count", "Minutes",
               "Subtotal Std €", "Subtotal Cust €"]
    ROW += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=ROW, column=i, value=h)
        _style(c, font=_font(bold=True, size=9, color="FFFFFF"), fill=SECTION_FILL, align="center")

    def data_row(row, label, unit, std, cust, qty, fg, fh):
        ws.row_dimensions[row].height = 32
        W(row, 1, label, font=LABEL_FONT,  fill=LABEL_FILL)
        W(row, 2, unit,  font=LABEL_FONT,  fill=LABEL_FILL, align="center")
        W(row, 3, std,   font=INPUT_FONT,  fill=INPUT_FILL, align="right",
          num_fmt=EUR if unit == "€" else "0.00")
        W(row, 4, cust,  font=INPUT_FONT,  fill=INPUT_FILL, align="right",
          num_fmt=EUR if unit == "€" else "0.00")
        if qty is not None:
            W(row, 5, qty, font=INPUT_FONT, fill=INPUT_FILL, align="right")
        for col, fml in [(7, fg), (8, fh)]:
            ws.cell(row=row, column=col).value = fml
            _style(ws.cell(row=row, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
                   align="right", num_fmt=EUR)

    R7 = ROW + 1
    data_row(R7, "SMD Components – Time per assembly", "Ct",
             p['smd_time_std'], p['smd_time_cust'], p['smd_qty'],
             f"=C{R7}*E{R7}/100", f"=D{R7}*E{R7}/100")

    R8 = R7 + 1
    ws.row_dimensions[R8].height = 28
    for col, val in [(1, "Soldering 1-sided / 2-sided  (enter 1 or 2)"),
                     (2, "Ct"), (3, p['soldering_sides']),
                     (4, p['soldering_sides']), (5, p['soldering_sides'])]:
        f = LABEL_FONT if col in [1, 2] else INPUT_FONT
        fi = LABEL_FILL if col in [1, 2] else INPUT_FILL
        W(R8, col, val, font=f, fill=fi, align="center" if col == 2 else "right")
    R13 = R8 + 5
    for col, fml in [(7, f"=IF(E{R8}=2,C{R8}/E{R13}/100,0)"),
                     (8, f"=IF(E{R8}=2,D{R8}/E{R13}/100,0)")]:
        ws.cell(row=R8, column=col).value = fml
        _style(ws.cell(row=R8, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    R9 = R8 + 1
    ws.row_dimensions[R9].height = 24
    W(R9, 1, "FIX Printer per side", font=LABEL_FONT, fill=LABEL_FILL)
    W(R9, 3, p['fix_printer_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    W(R9, 4, p['fix_printer_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)

    R10 = R9 + 1
    ws.row_dimensions[R10].height = 24
    W(R10, 1, "FIX Machine per side", font=LABEL_FONT, fill=LABEL_FILL)
    W(R10, 3, p['fix_machine_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    W(R10, 4, p['fix_machine_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)

    R11 = R10 + 1
    ws.row_dimensions[R11].height = 36
    W(R11, 1, "SMD Feeders – Setup cost", font=LABEL_FONT, fill=LABEL_FILL)
    W(R11, 2, "€",                        font=LABEL_FONT, fill=LABEL_FILL, align="center")
    W(R11, 3, p['feeder_cost_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    W(R11, 4, p['feeder_cost_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    W(R11, 5, p['feeder_qty'],       font=INPUT_FONT, fill=INPUT_FILL, align="right")
    for col, fml in [
        (7, f"=(((C{R9}*E{R8})+(C{R10}*E{R8})+(C{R11}*E{R11}))/G3)"),
        (8, f"=(((D{R9}*E{R8})+(D{R10}*E{R8})+(D{R11}*E{R11}))/H3)")
    ]:
        ws.cell(row=R11, column=col).value = fml
        _style(ws.cell(row=R11, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    R12 = R11 + 1
    ws.row_dimensions[R12].height = 24
    W(R12, 1, "THT Component count", font=LABEL_FONT, fill=LABEL_FILL)
    W(R12, 5, p['tht_parts'], font=INPUT_FONT, fill=INPUT_FILL, align="right")

    ws.row_dimensions[R13].height = 24
    W(R13, 1, "Circuits per panel (Einzelschaltungen pro Nutzen)", font=LABEL_FONT, fill=LABEL_FILL)
    W(R13, 5, p['circuits_panel'], font=INPUT_FONT, fill=INPUT_FILL, align="right")

    R14 = R13 + 1
    ws.row_dimensions[R14].height = 24
    W(R14, 1, "FIX Selective or Wave solder", font=LABEL_FONT, fill=LABEL_FILL)
    W(R14, 2, "€", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    W(R14, 3, p['fix_selective_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    W(R14, 4, p['fix_selective_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)

    R15 = R14 + 1
    data_row(R15, "THT Components – Time", "Ct",
             p['tht_time_std'], p['tht_time_cust'], p['tht_qty'],
             f"=(((C{R14}*E{R8})/G3)+((C{R15}*E{R15})/100))",
             f"=(((D{R14}*E{R8})/H3)+((D{R15}*E{R15})/100))")

    R16 = R15 + 1
    ws.row_dimensions[R16].height = 28
    W(R16, 1, "Manual solder joints", font=LABEL_FONT, fill=LABEL_FILL)
    W(R16, 3, p['manual_joints_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right")
    W(R16, 4, p['manual_joints_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right")
    W(R16, 5, p['manual_time'],        font=INPUT_FONT, fill=INPUT_FILL, align="right")
    R18 = R16 + 2
    for col, fml in [(7, f"=C{R16}*C{R18}*E{R16}/60"), (8, f"=D{R16}*D{R18}*E{R16}/60")]:
        ws.cell(row=R16, column=col).value = fml
        _style(ws.cell(row=R16, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    R17 = R16 + 1
    ws.row_dimensions[R17].height = 24
    W(R17, 1, "Quality Control (QS) time [min]", font=LABEL_FONT, fill=LABEL_FILL)
    W(R17, 6, p['qs_time'], font=INPUT_FONT, fill=INPUT_FILL, align="right")
    for col, fml in [(7, f"=F{R17}*C{R18}/60"), (8, f"=F{R17}*D{R18}/60")]:
        ws.cell(row=R17, column=col).value = fml
        _style(ws.cell(row=R17, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    ws.row_dimensions[R18].height = 24
    W(R18, 1, "Hourly Rate (Faktor Stundensatz)", font=LABEL_FONT, fill=LABEL_FILL)
    W(R18, 2, "€", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    W(R18, 3, p['hourly_rate_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    W(R18, 4, p['hourly_rate_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)

    R19 = R18 + 1
    ws.row_dimensions[R19].height = 24
    W(R19, 1, "TOTAL LABOUR (without markup)", font=TOTAL_FONT, fill=TOTAL_FILL)
    for col, fml in [(7, f"=SUM(G{R7}:G{R18})"), (8, f"=SUM(H{R7}:H{R18})")]:
        ws.cell(row=R19, column=col).value = fml
        _style(ws.cell(row=R19, column=col), font=TOTAL_FONT, fill=TOTAL_FILL,
               align="right", num_fmt=EUR)

    # ── SECTION 2: Material ──────────────────────────────────────────────────
    ROW_MAT = R19 + 2
    ws.merge_cells(f"A{ROW_MAT}:H{ROW_MAT}")
    _style(ws.cell(row=ROW_MAT, column=1, value="SECTION 2 – Material"),
           font=SECTION_FONT, fill=SECTION_FILL)

    R25 = ROW_MAT + 1
    ws.row_dimensions[R25].height = 24
    W(R25, 1, "Material cost per unit", font=LABEL_FONT, fill=LABEL_FILL)
    W(R25, 2, "€", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    W(R25, 3, p['material_cost_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    W(R25, 4, p['material_cost_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    R26 = R25 + 1
    for col, fml in [(7, f"=(C{R25}*C{R26}/100)+C{R25}"),
                     (8, f"=(D{R25}*D{R26}/100)+D{R25}")]:
        ws.cell(row=R25, column=col).value = fml
        _style(ws.cell(row=R25, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    ws.row_dimensions[R26].height = 24
    W(R26, 1, "Material markup %", font=LABEL_FONT, fill=LABEL_FILL)
    W(R26, 3, p['material_markup_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00")
    W(R26, 4, p['material_markup_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00")

    R27 = R26 + 1
    ws.row_dimensions[R27].height = 24
    W(R27, 1, "Total material markup (on full order)", font=LABEL_FONT, fill=LABEL_FILL)
    W(R27, 2, "€", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    for col, fml in [(3, f"=C{R25}*C{R26}/100*G3"), (4, f"=D{R25}*D{R26}/100*H3")]:
        ws.cell(row=R27, column=col).value = fml
        _style(ws.cell(row=R27, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    # ── SECTION 3: Pricing ───────────────────────────────────────────────────
    ROW_PRICE = R27 + 2
    ws.merge_cells(f"A{ROW_PRICE}:H{ROW_PRICE}")
    _style(ws.cell(row=ROW_PRICE, column=1, value="SECTION 3 – Pricing"),
           font=SECTION_FONT, fill=SECTION_FILL)

    R30 = ROW_PRICE + 1
    ws.row_dimensions[R30].height = 22
    W(R30, 1, "Assembly price incl. Material (per unit)", font=LABEL_FONT, fill=LABEL_FILL)
    for col, fml in [(3, f"=G{R25}+G{R19}"), (4, f"=H{R25}+H{R19}")]:
        ws.cell(row=R30, column=col).value = fml
        _style(ws.cell(row=R30, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    R31 = R30 + 1
    ws.row_dimensions[R31].height = 22
    W(R31, 1, "Price + Skonto", font=LABEL_FONT, fill=LABEL_FILL)
    W(R31, 2, p['skonto'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00%")
    for col, fml in [(3, f"=C{R30}*B{R31}+C{R30}"), (4, f"=D{R30}*B{R31}+D{R30}")]:
        ws.cell(row=R31, column=col).value = fml
        _style(ws.cell(row=R31, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    R32 = R31 + 1
    ws.row_dimensions[R32].height = 22
    W(R32, 1, "Unit price incl. profit margin", font=TOTAL_FONT, fill=TOTAL_FILL)
    W(R32, 2, p['profit_margin'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00%")
    for col, fml in [(7, f"=C{R31}*B{R32}+C{R31}"), (8, f"=D{R31}*B{R32}+D{R31}")]:
        ws.cell(row=R32, column=col).value = fml
        _style(ws.cell(row=R32, column=col), font=TOTAL_FONT, fill=TOTAL_FILL,
               align="right", num_fmt=EUR)

    R33 = R32 + 1
    ws.row_dimensions[R33].height = 22
    W(R33, 1, "Panel price incl. skonto & profit margin", font=LABEL_FONT, fill=LABEL_FILL)
    for col, fml in [(3, f"=G{R32}*E{R13}"), (4, f"=H{R32}*E{R13}")]:
        ws.cell(row=R33, column=col).value = fml
        _style(ws.cell(row=R33, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    R34 = R33 + 1
    ws.row_dimensions[R34].height = 22
    W(R34, 1, "Total order value LABOUR + Material", font=TOTAL_FONT, fill=TOTAL_FILL)
    for col, fml in [(7, f"=G{R32}*G3"), (8, f"=H3*H{R32}")]:
        ws.cell(row=R34, column=col).value = fml
        _style(ws.cell(row=R34, column=col), font=TOTAL_FONT, fill=TOTAL_FILL,
               align="right", num_fmt=EUR)

    R35 = R34 + 1
    ws.row_dimensions[R35].height = 20
    W(R35, 1, "Order value LOHN only (excl. material)", font=LABEL_FONT, fill=LABEL_FILL)
    for col, fml in [(7, f"=G{R34}-(C{R25}*G3)"), (8, f"=H{R34}-(D{R25}*H3)")]:
        ws.cell(row=R35, column=col).value = fml
        _style(ws.cell(row=R35, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    # ── SECTION 4: Setup ─────────────────────────────────────────────────────
    ROW_SETUP = R35 + 2
    ws.merge_cells(f"A{ROW_SETUP}:H{ROW_SETUP}")
    _style(ws.cell(row=ROW_SETUP, column=1, value="SECTION 4 – Initial / Setup Costs"),
           font=SECTION_FONT, fill=SECTION_FILL)

    sh_headers = ["Description", "Unit", "Standard", "Customer", "Factor", "", "", "Customer Total"]
    ROW_SH = ROW_SETUP + 1
    for i, h in enumerate(sh_headers, 1):
        c = ws.cell(row=ROW_SH, column=i, value=h)
        _style(c, font=_font(bold=True, size=9, color="FFFFFF"), fill=SECTION_FILL, align="center")

    def setup_row(row, label, std, cust, formula_h, factor=None):
        ws.row_dimensions[row].height = 28
        W(row, 1, label, font=LABEL_FONT, fill=LABEL_FILL)
        W(row, 2, "€",   font=LABEL_FONT, fill=LABEL_FILL, align="center")
        W(row, 3, std,   font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
        W(row, 4, cust,  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
        if factor is not None:
            W(row, 5, factor, font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00")
        ws.cell(row=row, column=8).value = formula_h
        _style(ws.cell(row=row, column=8), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    R38 = ROW_SH + 1
    setup_row(R38, "SMD Assembly Setup per side (Valor, Programme, EFA, 3D-AOI)",
              p['setup_smd_std'], p['setup_smd_cust'], f"=D{R38}*E{R8}")
    R39 = R38 + 1
    setup_row(R39, "Setup Printer programme",
              p['setup_printer_std'], p['setup_printer_cust'], f"=D{R39}*E{R8}")
    R40 = R39 + 1
    setup_row(R40, "Setup SMD per part (VPL, SiplacePro, Feeder, 3D-AOI)",
              p['setup_smd_part_std'], p['setup_smd_part_cust'], f"=D{R40}*E{R11}")
    R41 = R40 + 1
    setup_row(R41, "Setup THT per part (Selective programme)",
              p['setup_tht_part_std'], p['setup_tht_part_cust'], f"=D{R41}*E{R15}")
    R42 = R41 + 1
    setup_row(R42, "Setup LP Supplier", 0, p['setup_lp_cust'], f"=D{R42}")
    R43 = R42 + 1
    setup_row(R43, "Stencil Top", p['stencil_top_std'], p['stencil_top_cust'],
              f"=D{R43}+(D{R43}*E{R43})", factor=p['stencil_top_factor'])
    R44 = R43 + 1
    setup_row(R44, "Stencil Bottom", p['stencil_bot_std'], p['stencil_bot_cust'],
              f"=D{R44}+(D{R44}*E{R44})", factor=p['stencil_bot_factor'])
    R45 = R44 + 1
    setup_row(R45, "Order processing",
              p['order_processing_std'], p['order_processing_cust'], f"=D{R45}")

    R46 = R45 + 1
    ws.row_dimensions[R46].height = 24
    W(R46, 1, "TOTAL Initial Costs", font=TOTAL_FONT, fill=TOTAL_FILL)
    ws.cell(row=R46, column=8).value = f"=SUM(H{R38}:H{R45})"
    _style(ws.cell(row=R46, column=8), font=TOTAL_FONT, fill=TOTAL_FILL,
           align="right", num_fmt=EUR)

    # ── SECTION 5: Project Summary ───────────────────────────────────────────
    ROW_PROJ = R46 + 2
    ws.merge_cells(f"A{ROW_PROJ}:H{ROW_PROJ}")
    _style(ws.cell(row=ROW_PROJ, column=1, value="SECTION 5 – Project Summary"),
           font=SECTION_FONT, fill=SECTION_FILL)

    summaries = [
        ("Project price serial production incl. setup costs",      f"=H{R46}+H{R34}", True),
        ("Project price incl. setup + prototype",                   f"=H{R46}+H{R34}+G{R34}", False),
        ("Project price serial production incl. setup (customer)",  f"=H{R34}+H{R46}", False),
    ]
    for i, (label, fml, is_total) in enumerate(summaries, 1):
        row = ROW_PROJ + i
        ws.row_dimensions[row].height = 24
        W(row, 1, label, font=TOTAL_FONT if is_total else LABEL_FONT,
          fill=TOTAL_FILL if is_total else LABEL_FILL)
        ws.cell(row=row, column=8).value = fml
        _style(ws.cell(row=row, column=8),
               font=TOTAL_FONT if is_total else FORMULA_FONT,
               fill=TOTAL_FILL if is_total else RESULT_FILL,
               align="right", num_fmt=EUR)

    # ── STD Days ─────────────────────────────────────────────────────────────
    R_STD = ROW_PROJ + len(summaries) + 2
    ws.row_dimensions[R_STD].height = 24
    W(R_STD, 1, "STD Lead Time & Cost per Hour/Day", font=LABEL_FONT, fill=LABEL_FILL)
    W(R_STD, 2, p['std_days'], font=INPUT_FONT, fill=INPUT_FILL, align="right")
    for col, fml in [(7, f"=G{R34}/B{R_STD}/8"), (8, f"=H{R34}/B{R_STD}/8")]:
        ws.cell(row=R_STD, column=col).value = fml
        _style(ws.cell(row=R_STD, column=col), font=FORMULA_FONT, fill=RESULT_FILL,
               align="right", num_fmt=EUR)

    # ── Footer ───────────────────────────────────────────────────────────────
    R_FOOT = R_STD + 2
    W(R_FOOT,     1, f"Created by: {p['created_by']}", font=_font(bold=True, size=9), fill=WHITE_FILL)
    W(R_FOOT + 1, 1, f"Date: {p['today']}",            font=_font(size=9),           fill=WHITE_FILL)

    ws.freeze_panes = "A7"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()