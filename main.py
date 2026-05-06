"""
PCB Assembly Cost Calculator
Translated from German Excel: "Baugruppe LOHN" (Assembly Labour Cost Sheet)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ─────────────────────────────────────────────
#  INPUTS
# ─────────────────────────────────────────────

def get_inputs():
    print("=" * 60)
    print("  PCB ASSEMBLY COST CALCULATOR")
    print("=" * 60)
    print()

    def ask(prompt, default, typ=float):
        val = input(f"  {prompt} [{default}]: ").strip()
        return typ(val) if val else typ(default)

    print("── General ──────────────────────────────────────────")
    quantity        = ask("Total quantity (H2)", 100, int)
    circuits_panel  = ask("Circuits per panel / Einzelschaltungen pro Nutzen (E13)", 4, int)

    print()
    print("── SMD Assembly ─────────────────────────────────────")
    smd_time_std    = ask("SMD time per assembly [Ct] – Standard (C7)", 5.5)
    smd_time_cust   = ask("SMD time per assembly [Ct] – Customer (D7)", 4.0)
    smd_qty         = ask("SMD component count (E7)", 120, int)

    soldering_sides = ask("Soldering sides 1 or 2 (E8)", 1, int)

    fix_printer_std = ask("FIX printer per side € – Standard (C9)", 30.0)
    fix_printer_cust= ask("FIX printer per side € – Customer (D9)", 30.0)
    fix_machine_std = ask("FIX machine per side € – Standard (C10)", 30.0)
    fix_machine_cust= ask("FIX machine per side € – Customer (D10)", 30.0)

    feeder_cost_std = ask("SMD feeder/setup cost € – Standard (C11)", 6.5)
    feeder_cost_cust= ask("SMD feeder/setup cost € – Customer (D11)", 6.5)
    feeder_qty      = ask("Number of SMD feeder types (E11)", 33, int)

    print()
    print("── THT Assembly ─────────────────────────────────────")
    tht_parts       = ask("THT component count (E12)", 0, int)
    fix_selective_std  = ask("FIX selective/wave solder € – Standard (C14)", 30.0)
    fix_selective_cust = ask("FIX selective/wave solder € – Customer (D14)", 30.0)
    tht_time_std    = ask("THT time [Ct] – Standard (C15)", 55.0)
    tht_time_cust   = ask("THT time [Ct] – Customer (D15)", 45.0)
    tht_qty         = ask("THT component count for time calc (E15)", 0, int)

    print()
    print("── Manual Soldering ─────────────────────────────────")
    manual_joints_std  = ask("Manual solder joints – Standard (C16)", 0, int)
    manual_joints_cust = ask("Manual solder joints – Customer (D16)", 0, int)
    manual_time     = ask("Time per joint [min] (E16) – leave 0 for none", 0.0)

    print()
    print("── QS / Hourly Rate ─────────────────────────────────")
    qs_time         = ask("QS time [min] (F17)", 1.8)
    hourly_rate_std = ask("Hourly rate € – Standard (C18)", 60.0)
    hourly_rate_cust= ask("Hourly rate € – Customer (D18)", 60.0)

    print()
    print("── Material ─────────────────────────────────────────")
    material_cost_std  = ask("Material cost per unit € – Standard (C25)", 0.0)
    material_cost_cust = ask("Material cost per unit € – Customer (D25)", 0.0)
    material_markup_std  = ask("Material markup % – Standard (C26)", 0.0)
    material_markup_cust = ask("Material markup % – Customer (D26)", 0.0)

    print()
    print("── Pricing ──────────────────────────────────────────")
    skonto          = ask("Skonto / early-payment discount (B31, e.g. 0 = 0%)", 0.0)
    profit_margin   = ask("Profit margin (B32, e.g. 0.05 = 5%)", 0.05)

    print()
    print("── Setup / Initial Costs ────────────────────────────")
    setup_smd_std   = ask("Setup SMD assembly per side € – Standard (C38)", 150.0)
    setup_smd_cust  = ask("Setup SMD assembly per side € – Customer (D38)", 90.0)
    setup_printer_std  = ask("Setup printer programme € – Standard (C39)", 45.0)
    setup_printer_cust = ask("Setup printer programme € – Customer (D39)", 45.0)
    setup_smd_part_std = ask("Setup SMD per part € – Standard (C40)", 6.75)
    setup_smd_part_cust= ask("Setup SMD per part € – Customer (D40)", 1.75)
    setup_tht_part_std = ask("Setup THT per part € – Standard (C41)", 6.0)
    setup_tht_part_cust= ask("Setup THT per part € – Customer (D41)", 2.5)
    setup_lp_cust   = ask("Setup LP supplier € – Customer (D42)", 0.0)
    stencil_top_std = ask("Stencil Top € – Standard (C43)", 150.0)
    stencil_top_cust= ask("Stencil Top € – Customer (D43)", 0.0)
    stencil_top_factor = ask("Stencil Top surcharge factor (E43, e.g. 0.2)", 0.2)
    stencil_bot_std = ask("Stencil Bottom € – Standard (C44)", 150.0)
    stencil_bot_cust= ask("Stencil Bottom € – Customer (D44)", 0.0)
    stencil_bot_factor = ask("Stencil Bottom surcharge factor (E44)", 0.2)
    order_processing_std  = ask("Order processing € – Standard (C45)", 60.0)
    order_processing_cust = ask("Order processing € – Customer (D45)", 60.0)

    print()
    print("── STD Days ─────────────────────────────────────────")
    std_days        = ask("Standard lead time days (B54)", 110, int)

    print()
    created_by = input("  Created by [Ashok Kumar]: ").strip() or "Ashok Kumar"
    today = date.today().strftime("%d.%m.%Y")

    return dict(
        quantity=quantity, circuits_panel=circuits_panel,
        smd_time_std=smd_time_std, smd_time_cust=smd_time_cust, smd_qty=smd_qty,
        soldering_sides=soldering_sides,
        fix_printer_std=fix_printer_std, fix_printer_cust=fix_printer_cust,
        fix_machine_std=fix_machine_std, fix_machine_cust=fix_machine_cust,
        feeder_cost_std=feeder_cost_std, feeder_cost_cust=feeder_cost_cust, feeder_qty=feeder_qty,
        tht_parts=tht_parts,
        fix_selective_std=fix_selective_std, fix_selective_cust=fix_selective_cust,
        tht_time_std=tht_time_std, tht_time_cust=tht_time_cust, tht_qty=tht_qty,
        manual_joints_std=manual_joints_std, manual_joints_cust=manual_joints_cust,
        manual_time=manual_time,
        qs_time=qs_time,
        hourly_rate_std=hourly_rate_std, hourly_rate_cust=hourly_rate_cust,
        material_cost_std=material_cost_std, material_cost_cust=material_cost_cust,
        material_markup_std=material_markup_std, material_markup_cust=material_markup_cust,
        skonto=skonto, profit_margin=profit_margin,
        setup_smd_std=setup_smd_std, setup_smd_cust=setup_smd_cust,
        setup_printer_std=setup_printer_std, setup_printer_cust=setup_printer_cust,
        setup_smd_part_std=setup_smd_part_std, setup_smd_part_cust=setup_smd_part_cust,
        setup_tht_part_std=setup_tht_part_std, setup_tht_part_cust=setup_tht_part_cust,
        setup_lp_cust=setup_lp_cust,
        stencil_top_std=stencil_top_std, stencil_top_cust=stencil_top_cust,
        stencil_top_factor=stencil_top_factor,
        stencil_bot_std=stencil_bot_std, stencil_bot_cust=stencil_bot_cust,
        stencil_bot_factor=stencil_bot_factor,
        order_processing_std=order_processing_std, order_processing_cust=order_processing_cust,
        std_days=std_days, created_by=created_by, today=today,
    )


# ─────────────────────────────────────────────
#  STYLES
# ─────────────────────────────────────────────

def _font(bold=False, size=10, color="000000", name="Arial"):
    return Font(name=name, size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL   = _fill("1F4E79")
SECTION_FILL  = _fill("2E75B6")
INPUT_FILL    = _fill("E2EFDA")
LABEL_FILL    = _fill("D6DCE4")
RESULT_FILL   = _fill("FFF2CC")
TOTAL_FILL    = _fill("FFD966")
WHITE_FILL    = _fill("FFFFFF")

HEADER_FONT   = _font(bold=True, size=11, color="FFFFFF")
SECTION_FONT  = _font(bold=True, size=10, color="FFFFFF")
LABEL_FONT    = _font(bold=False, size=10)
INPUT_FONT    = _font(bold=False, size=10, color="0000FF")   # blue = user input
FORMULA_FONT  = _font(bold=False, size=10, color="000000")
TOTAL_FONT    = _font(bold=True,  size=10)
RESULT_FONT   = _font(bold=True,  size=10, color="1F4E79")


def style(cell, font=None, fill=None, align="left", bold=None, border=True, num_fmt=None):
    if font:  cell.font = font
    if bold is not None:
        f = cell.font.copy()
        cell.font = Font(name=f.name, size=f.size, bold=bold, color=f.color.rgb if f.color else "000000")
    if fill:  cell.fill = fill
    if border: cell.border = _border()
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if num_fmt: cell.number_format = num_fmt


# ─────────────────────────────────────────────
#  EXCEL GENERATION
# ─────────────────────────────────────────────

def build_excel(p, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{p['quantity']} pcs – {p['today']}"

    # Column widths
    ws.column_dimensions['A'].width = 58
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16

    EUR = '#,##0.00 €'
    PCT = '0.00%'

    def w(row, col, val, font=None, fill=None, align="left", num_fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        style(c, font=font, fill=fill, align=align, num_fmt=num_fmt)
        return c

    # ── Title row ──
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value=f"PCB ASSEMBLY COST SHEET  |  Qty: {p['quantity']}  |  Date: {p['today']}")
    style(c, font=HEADER_FONT, fill=HEADER_FILL, align="center")

    # ── Legend row ──
    ws.row_dimensions[2].height = 18
    w(2, 1, "Legend:", font=_font(bold=True, size=9))
    w(2, 3, "Standard", font=_font(bold=True, size=9, color="FFFFFF"), fill=SECTION_FILL, align="center")
    w(2, 4, "Customer", font=_font(bold=True, size=9, color="FFFFFF"), fill=_fill("375623"), align="center")
    w(2, 7, "Subtotal Standard", font=_font(bold=True, size=9, color="FFFFFF"), fill=SECTION_FILL, align="center")
    w(2, 8, "Subtotal Customer", font=_font(bold=True, size=9, color="FFFFFF"), fill=_fill("375623"), align="center")

    # ── Quantity row ──
    ws.row_dimensions[3].height = 18
    w(3, 7, "Subtotal", font=SECTION_FONT, fill=SECTION_FILL, align="center")
    w(3, 8, "Subtotal", font=SECTION_FONT, fill=_fill("375623"), align="center")
    w(3, 1, "PCB:", font=_font(bold=True))
    ws.cell(row=3, column=7).value = p['quantity']
    ws.cell(row=3, column=8).value = p['quantity']
    for col in [7, 8]:
        style(ws.cell(row=3, column=col), font=INPUT_FONT, fill=INPUT_FILL, align="center")

    # ─── SECTION: Labour Assembly ───
    ROW = 5
    ws.row_dimensions[ROW].height = 20
    ws.merge_cells(f"A{ROW}:H{ROW}")
    c = ws.cell(row=ROW, column=1, value="SECTION 1 – Labour Assembly (LOHN)")
    style(c, font=SECTION_FONT, fill=SECTION_FILL)

    headers = ["Description", "Unit", "Standard", "Customer", "Qty / Count", "Minutes", "Subtotal Std €", "Subtotal Cust €"]
    ROW += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=ROW, column=i, value=h)
        style(c, font=_font(bold=True, size=9, color="FFFFFF"), fill=SECTION_FILL, align="center")

    # helper to write a data row
    def data_row(row, label, unit, std, cust, qty, formula_g, formula_h, fill_g=None, fill_h=None):
        ws.row_dimensions[row].height = 32
        w(row, 1, label, font=LABEL_FONT, fill=LABEL_FILL)
        w(row, 2, unit,  font=LABEL_FONT, fill=LABEL_FILL, align="center")
        c3 = w(row, 3, std,  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR if unit=="€" else "0.00")
        c4 = w(row, 4, cust, font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR if unit=="€" else "0.00")
        if qty is not None:
            w(row, 5, qty, font=INPUT_FONT, fill=INPUT_FILL, align="right")
        ws.cell(row=row, column=7).value = formula_g
        ws.cell(row=row, column=8).value = formula_h
        style(ws.cell(row=row, column=7), font=FORMULA_FONT, fill=fill_g or RESULT_FILL, align="right", num_fmt=EUR)
        style(ws.cell(row=row, column=8), font=FORMULA_FONT, fill=fill_h or _fill("E2EFDA"), align="right", num_fmt=EUR)

    # Map rows to Excel rows (starting at 7 like original)
    # R7: SMD Bauteile
    R7 = ROW + 1
    data_row(R7, "SMD Components – Time per assembly", "Ct",
             p['smd_time_std'], p['smd_time_cust'], p['smd_qty'],
             f"=C{R7}*E{R7}/100", f"=D{R7}*E{R7}/100")

    # R8: Soldering sides
    R8 = R7 + 1
    ws.row_dimensions[R8].height = 28
    w(R8, 1, "Soldering 1-sided / 2-sided  (enter 1 or 2)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R8, 2, "Ct", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    w(R8, 3, p['soldering_sides'], font=INPUT_FONT, fill=INPUT_FILL, align="right")
    w(R8, 4, p['soldering_sides'], font=INPUT_FONT, fill=INPUT_FILL, align="right")
    w(R8, 5, p['soldering_sides'], font=INPUT_FONT, fill=INPUT_FILL, align="right")
    # Soldering formula uses circuits_panel row
    R13 = R8 + 5   # E13 equivalent
    ws.cell(row=R8, column=7).value = f"=IF(E{R8}=2,C{R8}/E{R13}/100,0)"
    ws.cell(row=R8, column=8).value = f"=IF(E{R8}=2,D{R8}/E{R13}/100,0)"
    for col in [7, 8]:
        style(ws.cell(row=R8, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    # R9: FIX Printer
    R9 = R8 + 1
    ws.row_dimensions[R9].height = 24
    w(R9, 1, "FIX Printer per side (standard €30)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R9, 3, p['fix_printer_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    w(R9, 4, p['fix_printer_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)

    # R10: FIX Machine
    R10 = R9 + 1
    ws.row_dimensions[R10].height = 24
    w(R10, 1, "FIX Machine per side (standard €30)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R10, 3, p['fix_machine_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    w(R10, 4, p['fix_machine_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)

    # R11: Feeder/SMD Setup
    R11 = R10 + 1
    ws.row_dimensions[R11].height = 36
    w(R11, 1, "SMD Feeders – Setup cost (all components, Teaching, Stock)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R11, 2, "€",  font=LABEL_FONT, fill=LABEL_FILL, align="center")
    w(R11, 3, p['feeder_cost_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    w(R11, 4, p['feeder_cost_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    w(R11, 5, p['feeder_qty'],       font=INPUT_FONT, fill=INPUT_FILL, align="right")
    ws.cell(row=R11, column=7).value = f"=(((C{R9}*E{R8})+(C{R10}*E{R8})+(C{R11}*E{R11}))/G3)"
    ws.cell(row=R11, column=8).value = f"=(((D{R9}*E{R8})+(D{R10}*E{R8})+(D{R11}*E{R11}))/H3)"
    for col in [7, 8]:
        style(ws.cell(row=R11, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    # R12: THT parts count
    R12 = R11 + 1
    ws.row_dimensions[R12].height = 24
    w(R12, 1, "THT Component count", font=LABEL_FONT, fill=LABEL_FILL)
    w(R12, 5, p['tht_parts'], font=INPUT_FONT, fill=INPUT_FILL, align="right")

    # R13: Circuits per panel
    ws.row_dimensions[R13].height = 24
    w(R13, 1, "Circuits per panel (Einzelschaltungen pro Nutzen)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R13, 5, p['circuits_panel'], font=INPUT_FONT, fill=INPUT_FILL, align="right")

    # R14: FIX Selective/Wave
    R14 = R13 + 1
    ws.row_dimensions[R14].height = 24
    w(R14, 1, "FIX Selective or Wave solder (standard €30)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R14, 2, "€",  font=LABEL_FONT, fill=LABEL_FILL, align="center")
    w(R14, 3, p['fix_selective_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    w(R14, 4, p['fix_selective_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)

    # R15: THT time
    R15 = R14 + 1
    data_row(R15, "THT Components – Time (standard 35 Ct)", "Ct",
             p['tht_time_std'], p['tht_time_cust'], p['tht_qty'],
             f"=(((C{R14}*E{R8})/G3)+((C{R15}*E{R15})/100))",
             f"=(((D{R14}*E{R8})/H3)+((D{R15}*E{R15})/100))")

    # R16: Manual soldering
    R16 = R15 + 1
    ws.row_dimensions[R16].height = 28
    w(R16, 1, "Manual solder joints (0.08 min → 5 sec each)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R16, 3, p['manual_joints_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right")
    w(R16, 4, p['manual_joints_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right")
    w(R16, 5, p['manual_time'],        font=INPUT_FONT, fill=INPUT_FILL, align="right")
    R18 = R16 + 2
    ws.cell(row=R16, column=7).value = f"=C{R16}*C{R18}*E{R16}/60"
    ws.cell(row=R16, column=8).value = f"=D{R16}*D{R18}*E{R16}/60"
    for col in [7, 8]:
        style(ws.cell(row=R16, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    # R17: QS
    R17 = R16 + 1
    ws.row_dimensions[R17].height = 24
    w(R17, 1, "Quality Control (QS) time [min]", font=LABEL_FONT, fill=LABEL_FILL)
    w(R17, 6, p['qs_time'], font=INPUT_FONT, fill=INPUT_FILL, align="right")
    ws.cell(row=R17, column=7).value = f"=F{R17}*C{R18}/60"
    ws.cell(row=R17, column=8).value = f"=F{R17}*D{R18}/60"
    for col in [7, 8]:
        style(ws.cell(row=R17, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    # R18: Hourly rate
    ws.row_dimensions[R18].height = 24
    w(R18, 1, "Hourly Rate (Faktor Stundensatz)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R18, 2, "€", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    w(R18, 3, p['hourly_rate_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    w(R18, 4, p['hourly_rate_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)

    # R19: Total Labour
    R19 = R18 + 1
    ws.row_dimensions[R19].height = 24
    w(R19, 1, "TOTAL LABOUR (without markup)", font=TOTAL_FONT, fill=TOTAL_FILL)
    ws.cell(row=R19, column=7).value = f"=SUM(G{R7}:G{R18})"
    ws.cell(row=R19, column=8).value = f"=SUM(H{R7}:H{R18})"
    for col in [7, 8]:
        style(ws.cell(row=R19, column=col), font=TOTAL_FONT, fill=TOTAL_FILL, align="right", num_fmt=EUR)

    # ─── SECTION: Material ───
    ROW_MAT = R19 + 2
    ws.merge_cells(f"A{ROW_MAT}:H{ROW_MAT}")
    c = ws.cell(row=ROW_MAT, column=1, value="SECTION 2 – Material")
    style(c, font=SECTION_FONT, fill=SECTION_FILL)

    R25 = ROW_MAT + 1
    ws.row_dimensions[R25].height = 24
    w(R25, 1, "Material cost per unit", font=LABEL_FONT, fill=LABEL_FILL)
    w(R25, 2, "€", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    w(R25, 3, p['material_cost_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    w(R25, 4, p['material_cost_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
    R26 = R25 + 1
    ws.cell(row=R25, column=7).value = f"=(C{R25}*C{R26}/100)+C{R25}"
    ws.cell(row=R25, column=8).value = f"=(D{R25}*D{R26}/100)+D{R25}"
    for col in [7, 8]:
        style(ws.cell(row=R25, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    ws.row_dimensions[R26].height = 24
    w(R26, 1, "Material markup %", font=LABEL_FONT, fill=LABEL_FILL)
    w(R26, 3, p['material_markup_std'],  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00")
    w(R26, 4, p['material_markup_cust'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00")

    R27 = R26 + 1
    ws.row_dimensions[R27].height = 24
    w(R27, 1, "Total material markup (on full order)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R27, 2, "€", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    ws.cell(row=R27, column=3).value = f"=C{R25}*C{R26}/100*G3"
    ws.cell(row=R27, column=4).value = f"=D{R25}*D{R26}/100*H3"
    for col in [3, 4]:
        style(ws.cell(row=R27, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    # ─── SECTION: Pricing ───
    ROW_PRICE = R27 + 2
    ws.merge_cells(f"A{ROW_PRICE}:H{ROW_PRICE}")
    c = ws.cell(row=ROW_PRICE, column=1, value="SECTION 3 – Pricing")
    style(c, font=SECTION_FONT, fill=SECTION_FILL)

    R30 = ROW_PRICE + 1
    ws.row_dimensions[R30].height = 22
    w(R30, 1, "Assembly price incl. Material (per unit)", font=LABEL_FONT, fill=LABEL_FILL)
    ws.cell(row=R30, column=3).value = f"=G{R25}+G{R19}"
    ws.cell(row=R30, column=4).value = f"=H{R25}+H{R19}"
    for col in [3, 4]:
        style(ws.cell(row=R30, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    R31 = R30 + 1
    ws.row_dimensions[R31].height = 22
    w(R31, 1, "Price + Skonto (early payment discount)", font=LABEL_FONT, fill=LABEL_FILL)
    w(R31, 2, p['skonto'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00%")
    ws.cell(row=R31, column=3).value = f"=C{R30}*B{R31}+C{R30}"
    ws.cell(row=R31, column=4).value = f"=D{R30}*B{R31}+D{R30}"
    for col in [3, 4]:
        style(ws.cell(row=R31, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    R32 = R31 + 1
    ws.row_dimensions[R32].height = 22
    w(R32, 1, "Unit price incl. profit margin", font=TOTAL_FONT, fill=TOTAL_FILL)
    w(R32, 2, p['profit_margin'], font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00%")
    ws.cell(row=R32, column=7).value = f"=C{R31}*B{R32}+C{R31}"
    ws.cell(row=R32, column=8).value = f"=D{R31}*B{R32}+D{R31}"
    for col in [7, 8]:
        style(ws.cell(row=R32, column=col), font=TOTAL_FONT, fill=TOTAL_FILL, align="right", num_fmt=EUR)

    R33 = R32 + 1
    ws.row_dimensions[R33].height = 22
    w(R33, 1, "Panel price incl. skonto & profit margin", font=LABEL_FONT, fill=LABEL_FILL)
    ws.cell(row=R33, column=3).value = f"=G{R32}*E{R13}"
    ws.cell(row=R33, column=4).value = f"=H{R32}*E{R13}"
    for col in [3, 4]:
        style(ws.cell(row=R33, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    R34 = R33 + 1
    ws.row_dimensions[R34].height = 22
    w(R34, 1, "Total order value LABOUR + Material", font=TOTAL_FONT, fill=TOTAL_FILL)
    ws.cell(row=R34, column=7).value = f"=G{R32}*G3"
    ws.cell(row=R34, column=8).value = f"=H3*H{R32}"
    for col in [7, 8]:
        style(ws.cell(row=R34, column=col), font=TOTAL_FONT, fill=TOTAL_FILL, align="right", num_fmt=EUR)

    R35 = R34 + 1
    ws.row_dimensions[R35].height = 20
    w(R35, 1, "Order value LOHN only (excl. material)", font=LABEL_FONT, fill=LABEL_FILL)
    ws.cell(row=R35, column=7).value = f"=G{R34}-(C{R25}*G3)"
    ws.cell(row=R35, column=8).value = f"=H{R34}-(D{R25}*H3)"
    for col in [7, 8]:
        style(ws.cell(row=R35, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    # ─── SECTION: Setup / Initial Costs ───
    ROW_SETUP = R35 + 2
    ws.merge_cells(f"A{ROW_SETUP}:H{ROW_SETUP}")
    c = ws.cell(row=ROW_SETUP, column=1, value="SECTION 4 – Initial / Setup Costs")
    style(c, font=SECTION_FONT, fill=SECTION_FILL)

    setup_header = ["Description", "Unit", "Standard", "Customer", "Factor", "", "", "Customer Total"]
    ROW_SH = ROW_SETUP + 1
    for i, h in enumerate(setup_header, 1):
        c = ws.cell(row=ROW_SH, column=i, value=h)
        style(c, font=_font(bold=True, size=9, color="FFFFFF"), fill=SECTION_FILL, align="center")

    def setup_row(row, label, std, cust, formula_h, factor=None):
        ws.row_dimensions[row].height = 28
        w(row, 1, label, font=LABEL_FONT, fill=LABEL_FILL)
        w(row, 2, "€",   font=LABEL_FONT, fill=LABEL_FILL, align="center")
        w(row, 3, std,   font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
        w(row, 4, cust,  font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt=EUR)
        if factor is not None:
            w(row, 5, factor, font=INPUT_FONT, fill=INPUT_FILL, align="right", num_fmt="0.00")
        ws.cell(row=row, column=8).value = formula_h
        style(ws.cell(row=row, column=8), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    R38 = ROW_SH + 1
    setup_row(R38, "SMD Assembly Setup per side (Valor, Programme, EFA, 3D-AOI)",
              p['setup_smd_std'], p['setup_smd_cust'], f"=D{R38}*E{R8}")
    R39 = R38 + 1
    setup_row(R39, "Setup Printer programme",
              p['setup_printer_std'], p['setup_printer_cust'], f"=D{R39}*E{R8}")
    R40 = R39 + 1
    setup_row(R40, "Setup SMD per part (VPL, SiplacePro, Feeder, 3D-AOI)",
              p['setup_smd_part_std'], p['setup_smd_part_cust'], f"=D{R40}*E{R11}")
    R41 = R40 + 1
    setup_row(R41, "Setup THT per part (Selective programme)",
              p['setup_tht_part_std'], p['setup_tht_part_cust'], f"=D{R41}*E{R15}")
    R42 = R41 + 1
    setup_row(R42, "Setup LP Supplier", 0, p['setup_lp_cust'], f"=D{R42}")
    R43 = R42 + 1
    setup_row(R43, "Stencil Top", p['stencil_top_std'], p['stencil_top_cust'],
              f"=D{R43}+(D{R43}*E{R43})", factor=p['stencil_top_factor'])
    R44 = R43 + 1
    setup_row(R44, "Stencil Bottom", p['stencil_bot_std'], p['stencil_bot_cust'],
              f"=D{R44}+(D{R44}*E{R44})", factor=p['stencil_bot_factor'])
    R45 = R44 + 1
    setup_row(R45, "Order processing (work order, order confirmation, invoice label)",
              p['order_processing_std'], p['order_processing_cust'], f"=D{R45}")

    R46 = R45 + 1
    ws.row_dimensions[R46].height = 24
    w(R46, 1, "TOTAL Initial Costs", font=TOTAL_FONT, fill=TOTAL_FILL)
    ws.cell(row=R46, column=8).value = f"=SUM(H{R38}:H{R45})"
    style(ws.cell(row=R46, column=8), font=TOTAL_FONT, fill=TOTAL_FILL, align="right", num_fmt=EUR)

    # ─── SECTION: Project Summary ───
    ROW_PROJ = R46 + 2
    ws.merge_cells(f"A{ROW_PROJ}:H{ROW_PROJ}")
    c = ws.cell(row=ROW_PROJ, column=1, value="SECTION 5 – Project Summary")
    style(c, font=SECTION_FONT, fill=SECTION_FILL)

    summaries = [
        ("Project price serial production incl. setup costs",           f"=H{R46}+H{R34}"),
        ("Project price incl. setup + prototype",                       f"=H{R46}+H{R34}+G{R34}"),
        ("Project price incl. setup for quantity (customer)",           f"=H{R34}+H{R46}"),
        ("Project price serial production incl. setup (customer)",      f"=H{R34}+H{R46}"),
    ]

    for i, (label, formula) in enumerate(summaries, 1):
        row = ROW_PROJ + i
        ws.row_dimensions[row].height = 24
        w(row, 1, label, font=TOTAL_FONT if i == 1 else LABEL_FONT, fill=TOTAL_FILL if i == 1 else LABEL_FILL)
        ws.cell(row=row, column=8).value = formula
        style(ws.cell(row=row, column=8), font=TOTAL_FONT if i == 1 else FORMULA_FONT,
              fill=TOTAL_FILL if i == 1 else RESULT_FILL, align="right", num_fmt=EUR)

    # ─── STD & Days ───
    R_STD = ROW_PROJ + len(summaries) + 2
    ws.row_dimensions[R_STD].height = 24
    w(R_STD, 1, "STD Lead Time & Cost per Hour/Day", font=LABEL_FONT, fill=LABEL_FILL)
    w(R_STD, 2, p['std_days'], font=INPUT_FONT, fill=INPUT_FILL, align="right")
    w(R_STD, 3, "€", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    w(R_STD, 4, "€", font=LABEL_FONT, fill=LABEL_FILL, align="center")
    ws.cell(row=R_STD, column=7).value = f"=G{R34}/B{R_STD}/8"
    ws.cell(row=R_STD, column=8).value = f"=H{R34}/B{R_STD}/8"
    for col in [7, 8]:
        style(ws.cell(row=R_STD, column=col), font=FORMULA_FONT, fill=RESULT_FILL, align="right", num_fmt=EUR)

    # ─── Footer ───
    R_FOOT = R_STD + 2
    w(R_FOOT,     1, f"Created by: {p['created_by']}", font=_font(bold=True, size=9), fill=WHITE_FILL)
    w(R_FOOT + 1, 1, f"Date: {p['today']}",            font=_font(size=9),           fill=WHITE_FILL)

    # ─── Freeze panes ───
    ws.freeze_panes = "A7"

    wb.save(output_path)
    print(f"\n✅  Excel saved → {output_path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    params = get_inputs()
    out = f"Assembly_Cost_{params['quantity']}pcs_{params['today'].replace('.', '')}.xlsx"
    build_excel(params, out)